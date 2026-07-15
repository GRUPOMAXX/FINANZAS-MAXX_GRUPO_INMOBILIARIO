from pathlib import Path
import sys
import shutil
import time
from copy import copy

from openpyxl import load_workbook


# ============================================================
# 0. RUTAS BASE
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent

if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))


# ============================================================
# 1. IMPORTAR REPORTES
# ============================================================

import Reporte_Cobranzas as reporte_cobranzas
import Reporte_clientes as reporte_gerencia


# ============================================================
# 2. VARIABLES
# ============================================================

ACTUALIZAR_BASE = True

NOMBRE_REPORTE_CIERRE = "Reporte_Cierre_Mes.xlsx"

BORRAR_REPORTES_INDIVIDUALES = True


# ============================================================
# 3. VARIABLES INTERNAS
# ============================================================

SALIDA_LISTA = reporte_cobranzas.SALIDA_LISTA
RUTA_REPORTE_CIERRE = SALIDA_LISTA / NOMBRE_REPORTE_CIERRE


# ============================================================
# 4. PREPARAR BASE UNA SOLA VEZ
# ============================================================

def preparar_base_unica():
    print("====================================")
    print("PREPARANDO BASE ÚNICA DE COBRANZAS")
    print("====================================")

    resultado = reporte_cobranzas.preparar_base_cobranzas(
        actualizar_base=ACTUALIZAR_BASE
    )

    if not isinstance(resultado, tuple) or len(resultado) < 2:
        raise ValueError(
            "❌ preparar_base_cobranzas no devolvió base y ruta. "
            "Revisa Reporte_Cobranzas.py."
        )

    base_ajustada = resultado[0]
    ruta_base = Path(resultado[1])

    print("✅ Base preparada correctamente:")
    print(f"   {ruta_base}")

    return base_ajustada, ruta_base


# ============================================================
# 5. GENERAR REPORTES INDIVIDUALES
# ============================================================

def generar_reporte_cobranzas_sin_descarga(base_ajustada):
    print("====================================")
    print("GENERANDO REPORTE DE COBRANZAS")
    print("====================================")

    if not hasattr(reporte_cobranzas, "generar_solo_reporte_cobranzas"):
        raise AttributeError(
            "❌ No existe generar_solo_reporte_cobranzas en Reporte_Cobranzas.py."
        )

    ruta_reporte = reporte_cobranzas.generar_solo_reporte_cobranzas(base_ajustada)
    ruta_reporte = Path(ruta_reporte)

    if not ruta_reporte.exists():
        raise FileNotFoundError(
            f"❌ No se generó Reporte_Cobranzas.xlsx:\n{ruta_reporte}"
        )

    print("✅ Reporte de cobranzas generado:")
    print(f"   {ruta_reporte}")

    return ruta_reporte


def generar_reporte_gerencia_sin_descarga(base_ajustada):
    print("====================================")
    print("GENERANDO REPORTE PARA GERENCIA")
    print("====================================")

    if hasattr(reporte_gerencia, "generar_reporte_gerencia"):
        ruta_reporte = reporte_gerencia.generar_reporte_gerencia(base_ajustada)

    elif hasattr(reporte_gerencia, "generar_solo_reporte_gerencia"):
        ruta_reporte = reporte_gerencia.generar_solo_reporte_gerencia(base_ajustada)

    elif hasattr(reporte_gerencia, "ejecutar_reporte_gerencia"):
        resultado = reporte_gerencia.ejecutar_reporte_gerencia(
            actualizar_base=False
        )

        if isinstance(resultado, tuple):
            ruta_reporte = resultado[-1]
        else:
            ruta_reporte = resultado

    else:
        raise AttributeError(
            "❌ No encontré función válida en Reporte_Cobranzas_Gerencia.py.\n"
            "Debe existir una de estas funciones:\n"
            "- generar_reporte_gerencia(base)\n"
            "- generar_solo_reporte_gerencia(base)\n"
            "- ejecutar_reporte_gerencia(actualizar_base=False)"
        )

    if ruta_reporte is None:
        ruta_reporte = SALIDA_LISTA / "Reporte_para_Gerencia.xlsx"

    ruta_reporte = Path(ruta_reporte)

    if not ruta_reporte.exists():
        raise FileNotFoundError(
            f"❌ No se generó Reporte_para_Gerencia.xlsx:\n{ruta_reporte}"
        )

    print("✅ Reporte para gerencia generado:")
    print(f"   {ruta_reporte}")

    return ruta_reporte


def generar_reportes_cierre(base_ajustada):
    ruta_cobranzas = generar_reporte_cobranzas_sin_descarga(base_ajustada)
    ruta_gerencia = generar_reporte_gerencia_sin_descarga(base_ajustada)

    return ruta_cobranzas, ruta_gerencia


# ============================================================
# 6. FUNCIONES AUXILIARES
# ============================================================

def eliminar_archivo_si_existe(ruta):
    ruta = Path(ruta)

    if not ruta.exists():
        return

    for intento in range(1, 8):
        try:
            ruta.unlink()
            return
        except PermissionError:
            print(f"⚠️ Archivo abierto. Reintento {intento}/7...")
            time.sleep(2)

    raise PermissionError(
        f"❌ No se puede reemplazar porque está abierto:\n{ruta}\n\n"
        "Cierra ese archivo y vuelve a ejecutar."
    )


def nombre_hoja_valido(nombre):
    nombre = str(nombre).strip()

    for caracter in ["[", "]", "*", "?", "/", "\\", ":"]:
        nombre = nombre.replace(caracter, "-")

    nombre = nombre[:31]

    return nombre if nombre else "Hoja"


def obtener_nombre_hoja_unico(wb, nombre_base):
    nombre_base = nombre_hoja_valido(nombre_base)

    if nombre_base not in wb.sheetnames:
        return nombre_base

    contador = 2

    while True:
        sufijo = f"_{contador}"
        nombre = nombre_base[:31 - len(sufijo)] + sufijo

        if nombre not in wb.sheetnames:
            return nombre

        contador += 1


def copiar_hoja_entre_libros(ws_origen, wb_destino, nombre_destino):
    """
    Copia una hoja visible de un workbook a otro workbook usando openpyxl.

    Copia:
    - valores y fórmulas
    - estilos
    - formatos numéricos
    - anchos de columnas
    - altos de filas
    - celdas combinadas
    - congelamiento de paneles
    - configuración de impresión
    - filtros
    - validaciones
    - formato condicional básico
    """

    ws_destino = wb_destino.create_sheet(title=nombre_destino)

    # Propiedades generales
    ws_destino.sheet_format = copy(ws_origen.sheet_format)
    ws_destino.sheet_properties = copy(ws_origen.sheet_properties)
    ws_destino.page_margins = copy(ws_origen.page_margins)
    ws_destino.page_setup = copy(ws_origen.page_setup)
    ws_destino.print_options = copy(ws_origen.print_options)
    ws_destino.freeze_panes = ws_origen.freeze_panes
    ws_destino.sheet_view.showGridLines = ws_origen.sheet_view.showGridLines

    # Anchos de columnas
    for col_letter, col_dim in ws_origen.column_dimensions.items():
        ws_destino.column_dimensions[col_letter].width = col_dim.width
        ws_destino.column_dimensions[col_letter].hidden = col_dim.hidden
        ws_destino.column_dimensions[col_letter].outlineLevel = col_dim.outlineLevel

    # Altos de filas
    for row_idx, row_dim in ws_origen.row_dimensions.items():
        ws_destino.row_dimensions[row_idx].height = row_dim.height
        ws_destino.row_dimensions[row_idx].hidden = row_dim.hidden
        ws_destino.row_dimensions[row_idx].outlineLevel = row_dim.outlineLevel

    # Celdas
    for fila in ws_origen.iter_rows():
        for celda_origen in fila:
            celda_destino = ws_destino[celda_origen.coordinate]

            celda_destino.value = celda_origen.value

            if celda_origen.has_style:
                celda_destino._style = copy(celda_origen._style)

            if celda_origen.number_format:
                celda_destino.number_format = celda_origen.number_format

            if celda_origen.font:
                celda_destino.font = copy(celda_origen.font)

            if celda_origen.fill:
                celda_destino.fill = copy(celda_origen.fill)

            if celda_origen.border:
                celda_destino.border = copy(celda_origen.border)

            if celda_origen.alignment:
                celda_destino.alignment = copy(celda_origen.alignment)

            if celda_origen.protection:
                celda_destino.protection = copy(celda_origen.protection)

            if celda_origen.hyperlink:
                celda_destino._hyperlink = copy(celda_origen.hyperlink)

            if celda_origen.comment:
                celda_destino.comment = copy(celda_origen.comment)

    # Celdas combinadas
    for rango in ws_origen.merged_cells.ranges:
        ws_destino.merge_cells(str(rango))

    # Autofiltro
    if ws_origen.auto_filter and ws_origen.auto_filter.ref:
        ws_destino.auto_filter.ref = ws_origen.auto_filter.ref

    # Validaciones
    try:
        ws_destino.data_validations = copy(ws_origen.data_validations)
    except Exception:
        pass

    # Formato condicional
    try:
        ws_destino.conditional_formatting = copy(ws_origen.conditional_formatting)
    except Exception:
        pass

    # Imágenes
    try:
        for imagen in ws_origen._images:
            ws_destino.add_image(copy(imagen), imagen.anchor)
    except Exception:
        pass

    # Gráficos
    try:
        for grafico in ws_origen._charts:
            ws_destino.add_chart(copy(grafico), grafico.anchor)
    except Exception:
        pass

    return ws_destino


def verificar_hojas_finales(ruta_final, hojas_esperadas):
    wb = load_workbook(ruta_final, read_only=True)

    faltantes = [
        hoja for hoja in hojas_esperadas
        if hoja not in wb.sheetnames
    ]

    wb.close()

    if faltantes:
        raise RuntimeError(
            "❌ Estas hojas no quedaron guardadas en Reporte_Cierre_Mes.xlsx:\n"
            + "\n".join(f"- {h}" for h in faltantes)
        )

    print("✅ Hojas copiadas y guardadas correctamente:")
    for hoja in hojas_esperadas:
        print(f"   - {hoja}")


def limpiar_reportes_individuales(ruta_cobranzas, ruta_gerencia, ruta_final):
    if not BORRAR_REPORTES_INDIVIDUALES:
        return

    print("====================================")
    print("LIMPIANDO REPORTES INDIVIDUALES")
    print("====================================")

    ruta_final = Path(ruta_final).resolve()

    rutas = [
        ("Reporte_Cobranzas", Path(ruta_cobranzas).resolve()),
        ("Reporte_para_Gerencia", Path(ruta_gerencia).resolve()),
    ]

    for nombre, ruta in rutas:
        if ruta == ruta_final:
            continue

        if not ruta.exists():
            continue

        try:
            ruta.unlink()
            print(f"🧹 Eliminado {nombre}: {ruta}")
        except PermissionError:
            print(f"⚠️ No se pudo eliminar porque está abierto: {ruta}")


# ============================================================
# 7. UNIR REPORTES SIN USAR EXCEL COM
# ============================================================

def unir_reportes_excel(ruta_cobranzas, ruta_gerencia, ruta_reporte_final):
    """
    Une reportes sin abrir Excel.

    1. Copia Reporte_Cobranzas.xlsx como Reporte_Cierre_Mes.xlsx.
    2. Abre Reporte_Cierre_Mes.xlsx con openpyxl.
    3. Abre Reporte_para_Gerencia.xlsx con openpyxl.
    4. Copia todas las hojas visibles de gerencia.
    5. Guarda el archivo final.
    6. Verifica que las hojas estén guardadas.
    """

    ruta_cobranzas = Path(ruta_cobranzas).resolve()
    ruta_gerencia = Path(ruta_gerencia).resolve()
    ruta_reporte_final = Path(ruta_reporte_final).resolve()

    if not ruta_cobranzas.exists():
        raise FileNotFoundError(
            f"❌ No existe Reporte_Cobranzas.xlsx:\n{ruta_cobranzas}"
        )

    if not ruta_gerencia.exists():
        raise FileNotFoundError(
            f"❌ No existe Reporte_para_Gerencia.xlsx:\n{ruta_gerencia}"
        )

    SALIDA_LISTA.mkdir(parents=True, exist_ok=True)

    eliminar_archivo_si_existe(ruta_reporte_final)

    shutil.copy2(ruta_cobranzas, ruta_reporte_final)

    print("====================================")
    print("CREANDO REPORTE_CIERRE_MES")
    print("====================================")
    print(f"Base:   {ruta_cobranzas}")
    print(f"Cierre: {ruta_reporte_final}")

    wb_final = load_workbook(ruta_reporte_final)
    wb_gerencia = load_workbook(ruta_gerencia)

    hojas_copiadas = []

    try:
        print("====================================")
        print("COPIANDO HOJAS VISIBLES DE GERENCIA")
        print("====================================")

        for ws_origen in wb_gerencia.worksheets:
            if ws_origen.sheet_state != "visible":
                continue

            nombre_origen = ws_origen.title
            nombre_destino = obtener_nombre_hoja_unico(wb_final, nombre_origen)

            print(f"📄 Copiando hoja visible: {nombre_origen}")

            copiar_hoja_entre_libros(
                ws_origen=ws_origen,
                wb_destino=wb_final,
                nombre_destino=nombre_destino,
            )

            hojas_copiadas.append(nombre_destino)

        if not hojas_copiadas:
            raise RuntimeError(
                "❌ No se encontró ninguna hoja visible en Reporte_para_Gerencia.xlsx."
            )

        wb_final.save(ruta_reporte_final)

    finally:
        wb_final.close()
        wb_gerencia.close()

    verificar_hojas_finales(
        ruta_final=ruta_reporte_final,
        hojas_esperadas=hojas_copiadas,
    )

    print("✅ Reporte_Cierre_Mes generado correctamente:")
    print(ruta_reporte_final)

    return ruta_reporte_final


# ============================================================
# 8. PROGRAMA PRINCIPAL
# ============================================================

def ejecutar_cierre_mes():
    base_ajustada, ruta_base = preparar_base_unica()

    ruta_cobranzas, ruta_gerencia = generar_reportes_cierre(base_ajustada)

    ruta_final = unir_reportes_excel(
        ruta_cobranzas=ruta_cobranzas,
        ruta_gerencia=ruta_gerencia,
        ruta_reporte_final=RUTA_REPORTE_CIERRE,
    )

    limpiar_reportes_individuales(
        ruta_cobranzas=ruta_cobranzas,
        ruta_gerencia=ruta_gerencia,
        ruta_final=ruta_final,
    )

    print("====================================")
    print("CIERRE DE MES TERMINADO")
    print("====================================")
    print("✅ Base utilizada:")
    print(f"   {ruta_base}")
    print("✅ Archivo final:")
    print(f"   {ruta_final}")

    return ruta_final


def main():
    try:
        ejecutar_cierre_mes()

    except Exception as e:
        print("====================================")
        print("❌ PROCESO DETENIDO POR ERROR")
        print("====================================")
        print(e)
        raise


if __name__ == "__main__":
    main()