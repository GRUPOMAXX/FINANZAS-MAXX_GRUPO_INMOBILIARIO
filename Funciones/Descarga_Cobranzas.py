# Programa para descargar la base de datos de cobranzas en Evolta

# Creado por Eduardo Miguel Huamani Acosta                            15/07/26

from pathlib import Path
from datetime import datetime, date
import time
import shutil
import os
import sys
import re
import unicodedata
import pandas as pd

from dotenv import load_dotenv

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BASE_DIR = Path(__file__).resolve().parent.parent

if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from Conexiones.connection import (
    FECHA_INICIO_REPORTE,
    FECHA_FIN_REPORTE,
)

try:
    from Conexiones.connection import FECHA_CORTE_REPORTE
except ImportError:
    FECHA_CORTE_REPORTE = None


# =========================================================
# 1. VARIABLES QUE PUEDES MODIFICAR
# =========================================================

LOGIN_URL = "https://v4.evolta.pe/Login/Acceso/Index"
REPORTE_COBRANZA_URL = "https://v4.evolta.pe/Reportes/ReporteCobranza/Index"

DOWNLOADS_DIR = Path(r"C:\Users\ehuamani\Downloads")

load_dotenv(BASE_DIR / ".env")
load_dotenv(Path(__file__).resolve().parent / ".env")

LOGIN_AUTOMATICO = True

USUARIO_EVOLTA = os.getenv("USUARIO_EVOLTA")
CLAVE_EVOLTA = os.getenv("CLAVE_EVOLTA")

ENTRADA_COBRANZAS = BASE_DIR / "Flujo" / "Input" / "Cobranzas"
ENTRADA_COBRANZAS.mkdir(parents=True, exist_ok=True)

SALIDA_LISTA = BASE_DIR / "Flujo" / "Output"
SALIDA_LISTA.mkdir(parents=True, exist_ok=True)

FORMATO_EVOLTA = "Csv"

SEPARADOR_CSV = ","

CONVERTIR_CSV_A_EXCEL = True

OVERWRITE_EXISTING = True

BORRAR_ORIGINAL_DOWNLOADS = False

TIEMPO_ESPERA_DESCARGA = 180

INTENTOS_DESCARGA = 2

NOMBRE_HOJA_BASE = "Base_Cobranzas"
NOMBRE_BASE_CONSOLIDADA = "Base_Cobranzas_Consolidada.xlsx"
NOMBRE_BASE_AJUSTADA = "Base_Cobranzas_Ajustado.xlsx"

COLUMNAS_FECHA = [
    "Fecha_Programada",
    "FechaPago",
    "FechaOperacionComercial"
]

COLUMNAS_NUMERO_FIJAS = [
    "Monto_Cuota",
    "Monto_Cuota_Pagado",
    "SaldoPorPagarCuota"
]

COLUMNAS_NUMERO_CON_INDICE = [
    "PrecioLista",
    "PrecioVenta"
]


# =========================================================
# 2. FUNCIONES DE FECHA
# =========================================================

def convertir_fecha(fecha_texto: str) -> date:
    return datetime.strptime(fecha_texto, "%d/%m/%Y").date()


def fecha_texto(fecha: date) -> str:
    return fecha.strftime("%d/%m/%Y")


def generar_rangos_por_anio(fecha_inicio_txt: str, fecha_fin_txt: str):
    inicio = convertir_fecha(fecha_inicio_txt)
    fin = convertir_fecha(fecha_fin_txt)

    if inicio > fin:
        raise ValueError("La fecha de inicio no puede ser mayor que la fecha de fin.")

    rangos = []

    for anio in range(inicio.year, fin.year + 1):
        inicio_anio = date(anio, 1, 1)
        fin_anio = date(anio, 12, 31)

        desde = max(inicio, inicio_anio)
        hasta = min(fin, fin_anio)

        rangos.append({
            "anio": anio,
            "fecha_inicio": fecha_texto(desde),
            "fecha_fin": fecha_texto(hasta)
        })

    return rangos


# =========================================================
# 3. FUNCIONES PARA ENTRAR A EVOLTA
# =========================================================

def obtener_xpath_exportar():
    return """
    //button[contains(normalize-space(.), 'Exportar')]
    | //a[contains(normalize-space(.), 'Exportar')]
    | //input[@value='Exportar']
    | //*[@title='Exportar']
    """


def esperar_elemento(driver, xpath, timeout=30):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, xpath))
    )


def esperar_pagina_reporte(driver, timeout=40):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.XPATH, obtener_xpath_exportar()))
    )


def ir_a_reporte_cobranzas(driver):
    driver.get(REPORTE_COBRANZA_URL)
    esperar_pagina_reporte(driver)


def iniciar_sesion_evolta(driver):
    print("Ingresando a Evolta automáticamente...")

    if not USUARIO_EVOLTA or not CLAVE_EVOLTA:
        raise ValueError(
            "No se encontraron USUARIO_EVOLTA o CLAVE_EVOLTA en el archivo .env."
        )

    driver.get(LOGIN_URL)

    xpath_usuario = """
    //input[@type='email']
    | //input[contains(translate(@name, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'usuario')]
    | //input[contains(translate(@id, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'usuario')]
    | //input[contains(translate(@placeholder, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'usuario')]
    | //input[contains(translate(@name, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'user')]
    | //input[contains(translate(@id, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'user')]
    | //input[not(@type='password') and not(@type='hidden')][1]
    """

    xpath_clave = """
    //input[@type='password']
    | //input[contains(translate(@name, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'clave')]
    | //input[contains(translate(@id, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'clave')]
    | //input[contains(translate(@placeholder, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'clave')]
    | //input[contains(translate(@name, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'password')]
    | //input[contains(translate(@id, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'password')]
    """

    xpath_boton_ingresar = """
    //button[contains(normalize-space(.), 'Ingresar')]
    | //button[contains(normalize-space(.), 'Entrar')]
    | //button[contains(normalize-space(.), 'Acceder')]
    | //input[@type='submit']
    | //button[@type='submit']
    """

    campo_usuario = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, xpath_usuario))
    )

    campo_clave = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, xpath_clave))
    )

    campo_usuario.click()
    campo_usuario.send_keys(Keys.CONTROL, "a")
    campo_usuario.send_keys(Keys.BACKSPACE)
    campo_usuario.send_keys(USUARIO_EVOLTA)

    campo_clave.click()
    campo_clave.send_keys(Keys.CONTROL, "a")
    campo_clave.send_keys(Keys.BACKSPACE)
    campo_clave.send_keys(CLAVE_EVOLTA)

    boton_ingresar = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, xpath_boton_ingresar))
    )

    try:
        boton_ingresar.click()
    except Exception:
        try:
            ActionChains(driver).move_to_element(boton_ingresar).pause(0.5).click().perform()
        except Exception:
            driver.execute_script("arguments[0].click();", boton_ingresar)

    time.sleep(3)

    ir_a_reporte_cobranzas(driver)

    print("Sesión iniciada correctamente.")


def seleccionar_formato(driver, formato: str):
    if formato not in ["Csv", "Excel"]:
        raise ValueError("El formato debe ser 'Csv' o 'Excel'.")

    xpath_radio = f"//*[normalize-space()='{formato}']/preceding::input[@type='radio'][1]"

    radio = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, xpath_radio))
    )

    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", radio)
    time.sleep(0.5)

    driver.execute_script("arguments[0].click();", radio)
    time.sleep(0.5)


def escribir_fecha(driver, nombre_campo: str, fecha: str):
    print(f"Escribiendo {nombre_campo}: {fecha}")

    xpath = f"//*[normalize-space()='{nombre_campo}']/following::input[1]"
    campo = esperar_elemento(driver, xpath, timeout=30)

    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", campo)
    time.sleep(0.5)

    campo.click()
    time.sleep(0.2)

    campo.send_keys(Keys.CONTROL, "a")
    campo.send_keys(Keys.BACKSPACE)
    campo.send_keys(fecha)
    campo.send_keys(Keys.TAB)

    driver.execute_script("""
        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
        arguments[0].blur();
    """, campo)

    time.sleep(1)


def click_exportar(driver):
    print("Buscando botón Exportar...")

    boton = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, obtener_xpath_exportar()))
    )

    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton)
    time.sleep(1)

    print("Haciendo clic en Exportar...")

    try:
        boton.click()
    except Exception:
        try:
            ActionChains(driver).move_to_element(boton).pause(0.5).click().perform()
        except Exception:
            driver.execute_script("arguments[0].click();", boton)

    time.sleep(3)


# =========================================================
# 4. CONTROL DE DESCARGAS
# =========================================================

def configurar_descargas_chrome(driver):
    ruta_descarga = str(DOWNLOADS_DIR.resolve())

    try:
        driver.execute_cdp_cmd(
            "Browser.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": ruta_descarga
            }
        )
    except Exception:
        driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {
                "behavior": "allow",
                "downloadPath": ruta_descarga
            }
        )


def obtener_archivos_downloads():
    extensiones = [".csv", ".xlsx", ".xls"]

    return {
        archivo.name
        for archivo in DOWNLOADS_DIR.iterdir()
        if archivo.is_file()
        and archivo.suffix.lower() in extensiones
    }


def esperar_nueva_descarga_downloads(archivos_antes, tiempo_inicio, timeout=180):
    print("⌛ Esperando archivo nuevo en Descargas...")

    tiempo_limite = time.time() + timeout
    ultimo_aviso = time.time()

    while time.time() < tiempo_limite:
        temporales_recientes = []

        for temporal in DOWNLOADS_DIR.glob("*.crdownload"):
            try:
                if temporal.stat().st_mtime >= tiempo_inicio - 2:
                    temporales_recientes.append(temporal)
            except Exception:
                pass

        if temporales_recientes:
            time.sleep(1)
            continue

        candidatos = []

        for archivo in DOWNLOADS_DIR.iterdir():
            if not archivo.is_file():
                continue

            if archivo.suffix.lower() not in [".csv", ".xlsx", ".xls"]:
                continue

            try:
                nombre_nuevo = archivo.name not in archivos_antes
                modificado_reciente = archivo.stat().st_mtime >= tiempo_inicio - 2
            except Exception:
                continue

            if nombre_nuevo or modificado_reciente:
                candidatos.append(archivo)

        if candidatos:
            archivo_nuevo = max(candidatos, key=lambda f: f.stat().st_mtime)

            try:
                tamaño_1 = archivo_nuevo.stat().st_size
                time.sleep(1)
                tamaño_2 = archivo_nuevo.stat().st_size

                if tamaño_1 == tamaño_2 and tamaño_2 > 0:
                    print("Archivo detectado en Descargas:")
                    print(archivo_nuevo.name)
                    return archivo_nuevo
            except Exception:
                pass

        if time.time() - ultimo_aviso >= 15:
            print("Todavía esperando descarga...")
            ultimo_aviso = time.time()

        time.sleep(1)

    raise TimeoutError(
        "No se detectó ningún archivo nuevo en Descargas. "
    )


def renombrar_en_downloads(archivo_descargado: Path, anio: int):
    extension = archivo_descargado.suffix.lower()

    if extension not in [".csv", ".xlsx", ".xls"]:
        raise ValueError(f"Extensión no esperada: {extension}")

    nuevo_nombre_downloads = DOWNLOADS_DIR / f"Cobranzas_{anio}{extension}"

    if nuevo_nombre_downloads.exists() and OVERWRITE_EXISTING:
        nuevo_nombre_downloads.unlink()

    if nuevo_nombre_downloads.exists() and not OVERWRITE_EXISTING:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nuevo_nombre_downloads = DOWNLOADS_DIR / f"Cobranzas_{anio}_{timestamp}{extension}"

    archivo_descargado.rename(nuevo_nombre_downloads)

    if not nuevo_nombre_downloads.exists():
        raise FileNotFoundError(f"No se renombró correctamente en Descargas: {nuevo_nombre_downloads}")

    return nuevo_nombre_downloads


def copiar_renombrado_a_cobranzas(archivo_renombrado: Path):
    destino = ENTRADA_COBRANZAS / archivo_renombrado.name

    if destino.exists() and OVERWRITE_EXISTING:
        destino.unlink()

    if destino.exists() and not OVERWRITE_EXISTING:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = ENTRADA_COBRANZAS / f"{archivo_renombrado.stem}_{timestamp}{archivo_renombrado.suffix}"

    shutil.copy2(archivo_renombrado, destino)

    if not destino.exists():
        raise FileNotFoundError(f"No se copió correctamente el archivo a: {destino}")

    if BORRAR_ORIGINAL_DOWNLOADS:
        archivo_renombrado.unlink()
        print("Archivo original eliminado de Descargas.")

    return destino


# =========================================================
# 5. CONVERSIÓN CSV A EXCEL
# =========================================================

def leer_csv_seguro(ruta_csv: Path) -> pd.DataFrame:
    codificaciones = ["utf-8-sig", "utf-8", "latin1", "cp1252"]

    for encoding in codificaciones:
        try:
            df = pd.read_csv(
                ruta_csv,
                sep=SEPARADOR_CSV,
                dtype=str,
                encoding=encoding,
                keep_default_na=False
            )

            return df

        except Exception:
            continue

    raise ValueError(f"No se pudo leer el CSV: {ruta_csv}")


def convertir_csv_a_excel_seguro(ruta_csv: Path):
    df = leer_csv_seguro(ruta_csv)

    ruta_excel = ruta_csv.with_suffix(".xlsx")

    if ruta_excel.exists() and OVERWRITE_EXISTING:
        ruta_excel.unlink()

    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Base_CSV", index=False)

        ws = writer.book["Base_CSV"]

        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = "@"

    print(f"✅ Excel generado: {ruta_csv.name}")

    return ruta_excel


# =========================================================
# 6. LIMPIEZA DE ARCHIVOS PREVIOS
# =========================================================

def limpiar_archivos_previos_anio(anio: int):
    if not OVERWRITE_EXISTING:
        return

    posibles = [
        DOWNLOADS_DIR / f"Cobranzas_{anio}.csv",
        DOWNLOADS_DIR / f"Cobranzas_{anio}.xlsx",
        DOWNLOADS_DIR / f"Cobranzas_{anio}.xls",
        ENTRADA_COBRANZAS / f"Cobranzas_{anio}.csv",
        ENTRADA_COBRANZAS / f"Cobranzas_{anio}.xlsx",
        ENTRADA_COBRANZAS / f"Cobranzas_{anio}.xls",
    ]

    for archivo in posibles:
        if archivo.exists():
            try:
                archivo.unlink()
            except Exception:
                pass




# =========================================================
# 7. CONSOLIDACIÓN DE BASE DE COBRANZAS
# =========================================================

def normalizar_texto(texto):
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )
    return texto


def obtener_anio_archivo(nombre_archivo):
    coincidencia = re.search(r"(20\d{2})", nombre_archivo)
    if coincidencia:
        return coincidencia.group(1)
    return ""


def leer_archivo_cobranzas(ruta_archivo: Path) -> pd.DataFrame:
    if ruta_archivo.suffix.lower() == ".csv":
        return leer_csv_seguro(ruta_archivo)

    if ruta_archivo.suffix.lower() in [".xlsx", ".xls"]:
        return pd.read_excel(
            ruta_archivo,
            dtype=str,
            keep_default_na=False
        )

    raise ValueError(f"Formato no permitido para consolidar: {ruta_archivo}")


def obtener_archivos_para_consolidar():
    # Si existen CSV, se consolidan los CSV para no duplicar con los XLSX convertidos.
    archivos_csv = sorted(ENTRADA_COBRANZAS.glob("Cobranzas_*.csv"))

    if archivos_csv:
        return archivos_csv

    archivos_excel = sorted(ENTRADA_COBRANZAS.glob("Cobranzas_*.xlsx"))

    if archivos_excel:
        return archivos_excel

    raise FileNotFoundError(
        f"❌ No se encontraron archivos Cobranzas_*.csv ni Cobranzas_*.xlsx en: {ENTRADA_COBRANZAS}"
    )


def obtener_columnas_union(archivos_cobranzas):
    columnas_union = []
    columnas_vistas = set()

    for archivo in archivos_cobranzas:
        df = leer_archivo_cobranzas(archivo)

        for columna in list(df.columns):
            if columna not in columnas_vistas:
                columnas_union.append(columna)
                columnas_vistas.add(columna)

    return columnas_union


def ordenar_columnas_cobranzas(columnas_union):
    campos_inmueble = [
        "Tipo_Construccion",
        "Nombre_Tipo_Construccion",
        "TipoInmueble",
        "NroInmueble",
        "PrecioLista",
        "PrecioVenta"
    ]

    columnas_auxiliares = [
        "Archivo_Origen",
        "Anio_Descarga",
        "Fecha_Inicio_Reporte",
        "Fecha_Fin_Reporte"
    ]

    indices = set()
    columnas_a_reordenar = set()

    for columna in columnas_union:
        for campo in campos_inmueble:
            patron = rf"^{campo}_(\d+)$"
            coincidencia = re.match(patron, columna)

            if coincidencia:
                indice = int(coincidencia.group(1))
                indices.add(indice)
                columnas_a_reordenar.add(columna)

    columnas_inmueble_ordenadas = []

    for indice in sorted(indices):
        for campo in campos_inmueble:
            columna = f"{campo}_{indice}"

            if columna in columnas_union:
                columnas_inmueble_ordenadas.append(columna)

    if "Nombres_Titular" in columnas_union:
        columnas_a_reordenar.add("Nombres_Titular")
        columnas_inmueble_ordenadas.append("Nombres_Titular")

    posiciones = [
        i
        for i, columna in enumerate(columnas_union)
        if columna in columnas_a_reordenar
    ]

    posicion_insertar = min(posiciones) if posiciones else len(columnas_union)

    columnas_sin_bloque = [
        columna
        for columna in columnas_union
        if columna not in columnas_a_reordenar
    ]

    columnas_reordenadas = (
        columnas_sin_bloque[:posicion_insertar]
        + columnas_inmueble_ordenadas
        + columnas_sin_bloque[posicion_insertar:]
    )

    columnas_finales = [
        columna
        for columna in columnas_reordenadas
        if columna not in columnas_auxiliares
    ]

    columnas_finales = columnas_finales + columnas_auxiliares

    return columnas_finales


def convertir_columna_numero(serie):
    serie = serie.astype(str).str.strip()

    serie = serie.str.replace("S/", "", regex=False)
    serie = serie.str.replace("US$", "", regex=False)
    serie = serie.str.replace("$", "", regex=False)
    serie = serie.str.replace(" ", "", regex=False)
    serie = serie.str.replace(",", "", regex=False)

    serie = serie.replace("", pd.NA)

    return pd.to_numeric(serie, errors="coerce")


def es_columna_numero(columna):
    if columna in COLUMNAS_NUMERO_FIJAS:
        return True

    for campo in COLUMNAS_NUMERO_CON_INDICE:
        patron = rf"^{campo}_(\d+)$"

        if re.match(patron, columna):
            return True

    return False


def convertir_tipos_base(base):
    for columna in COLUMNAS_FECHA:
        if columna in base.columns:
            base[columna] = pd.to_datetime(
                base[columna],
                errors="coerce",
                dayfirst=True
            )

    for columna in base.columns:
        if es_columna_numero(columna):
            base[columna] = convertir_columna_numero(base[columna])

    return base


def aplicar_formatos_base_consolidada(ruta_excel):
    from openpyxl import load_workbook

    wb = load_workbook(ruta_excel)
    ws = wb[NOMBRE_HOJA_BASE]

    encabezados = {}
    for celda in ws[1]:
        encabezados[celda.value] = celda.column

    for columna in COLUMNAS_FECHA:
        if columna in encabezados:
            col_idx = encabezados[columna]
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=col_idx).number_format = "dd/mm/yyyy"

    for columna in encabezados:
        if es_columna_numero(columna):
            col_idx = encabezados[columna]
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=col_idx).number_format = "#,##0.00"

    ws.freeze_panes = "A2"

    wb.save(ruta_excel)


def generar_base_cobranzas_consolidada():
    """
    Genera la base consolidada original sin ajustes.
    """
    archivos_cobranzas = obtener_archivos_para_consolidar()

    print("\n✅ Archivos que se van a consolidar:")

    for archivo in archivos_cobranzas:
        print("-", archivo.name)

    columnas_union = obtener_columnas_union(archivos_cobranzas)
    columnas_finales = ordenar_columnas_cobranzas(columnas_union)

    bases = []

    for archivo in archivos_cobranzas:
        df = leer_archivo_cobranzas(archivo)
        anio_archivo = obtener_anio_archivo(archivo.name)

        df["Archivo_Origen"] = archivo.name
        df["Anio_Descarga"] = anio_archivo
        df["Fecha_Inicio_Reporte"] = FECHA_INICIO_REPORTE
        df["Fecha_Fin_Reporte"] = FECHA_FIN_REPORTE

        for columna in columnas_finales:
            if columna not in df.columns:
                df[columna] = ""

        df = df[columnas_finales]
        bases.append(df)

        print(f"{archivo.name}: {len(df)} filas, {len(df.columns)} columnas")

    base_consolidada = pd.concat(bases, ignore_index=True)

    for columna in columnas_finales:
        if columna not in base_consolidada.columns:
            base_consolidada[columna] = ""

    base_consolidada = base_consolidada[columnas_finales]
    base_consolidada = convertir_tipos_base(base_consolidada)

    ruta_base = SALIDA_LISTA / NOMBRE_BASE_CONSOLIDADA

    with pd.ExcelWriter(ruta_base, engine="openpyxl") as writer:
        base_consolidada.to_excel(
            writer,
            sheet_name=NOMBRE_HOJA_BASE,
            index=False
        )

    aplicar_formatos_base_consolidada(ruta_base)

    print("\n✅ Base consolidada original generada correctamente:")
    print(ruta_base.resolve())

    return ruta_base



def obtener_columna_cliente_ajuste(base):
    """
    Define la columna de cliente para aplicar los anticipos.
    Se prioriza NroDocumento porque identifica mejor al cliente.
    Si no existe, se usa Nombres_Titular.
    """
    if "NroDocumento" in base.columns:
        return "NroDocumento"

    if "Nombres_Titular" in base.columns:
        return "Nombres_Titular"

    raise ValueError(
        "No se encontró una columna válida para identificar al cliente "
        "(NroDocumento o Nombres_Titular)."
    )


def obtener_columna_departamento_ajuste(base):
    """
    Define la columna de departamento/unidad para aplicar los anticipos.
    En la base de Evolta, el departamento principal está en NroInmueble_1.
    """
    if "NroInmueble_1" in base.columns:
        return "NroInmueble_1"

    if "NroInmueble" in base.columns:
        return "NroInmueble"

    raise ValueError(
        "No se encontró una columna válida para identificar el departamento "
        "(NroInmueble_1 o NroInmueble)."
    )



def obtener_fecha_cierre_ajuste():
    """
    Lee FECHA_CORTE_REPORTE desde Conexiones/connection.py.
    """
    if FECHA_CORTE_REPORTE is None or str(FECHA_CORTE_REPORTE).strip() == "":
        return None

    fecha_cierre = pd.to_datetime(
        FECHA_CORTE_REPORTE,
        errors="coerce",
        dayfirst=True
    )

    if pd.isna(fecha_cierre):
        raise ValueError(
            "FECHA_CORTE_REPORTE no tiene un formato válido. "
            "Usa un formato como '30/06/2026'."
        )

    return fecha_cierre.normalize()


def reclasificar_estado_por_fecha_cierre(base, fecha_cierre):
    """
    Clasifica el estado según la fecha de cierre solo para filas con saldo.

    Regla:
    - Si SaldoPorPagarCuota > 0 y Fecha_Programada <= fecha de cierre: VENCIDO.
    - Si SaldoPorPagarCuota > 0 y Fecha_Programada > fecha de cierre: PENDIENTE.

    No modifica Monto_Cuota_Pagado.
    """
    base_estado = base.copy()

    if fecha_cierre is None:
        return base_estado, 0, 0

    columnas_necesarias = [
        "Estado",
        "Fecha_Programada",
        "SaldoPorPagarCuota",
    ]

    for columna in columnas_necesarias:
        if columna not in base_estado.columns:
            raise ValueError(f"No existe la columna requerida para reclasificar estados: {columna}")

    base_estado["Fecha_Programada"] = pd.to_datetime(
        base_estado["Fecha_Programada"],
        errors="coerce",
        dayfirst=True
    )

    base_estado["SaldoPorPagarCuota"] = pd.to_numeric(
        base_estado["SaldoPorPagarCuota"],
        errors="coerce"
    ).fillna(0)

    condicion_con_saldo = (
        (base_estado["SaldoPorPagarCuota"] > 0)
        & (base_estado["Fecha_Programada"].notna())
    )

    condicion_vencido = condicion_con_saldo & (base_estado["Fecha_Programada"] <= fecha_cierre)
    condicion_pendiente = condicion_con_saldo & (base_estado["Fecha_Programada"] > fecha_cierre)

    cantidad_vencidos = int(condicion_vencido.sum())
    cantidad_pendientes = int(condicion_pendiente.sum())

    base_estado.loc[condicion_vencido, "Estado"] = "VENCIDO"
    base_estado.loc[condicion_pendiente, "Estado"] = "PENDIENTE"

    return base_estado, cantidad_vencidos, cantidad_pendientes


def aplicar_fecha_cierre_base_ajustada(base):
    """
    Aplica la fecha de cierre en Base_Cobranzas_Ajustado.xlsx.
    """
    fecha_cierre = obtener_fecha_cierre_ajuste()

    if fecha_cierre is None:
        print("ℹ️ No se aplicó fecha de cierre porque FECHA_CORTE_REPORTE está vacío.")
        return base.copy()

    columnas_necesarias = [
        "Estado",
        "Fecha_Programada",
        "FechaPago",
        "Monto_Cuota_Pagado",
        "SaldoPorPagarCuota",
    ]

    for columna in columnas_necesarias:
        if columna not in base.columns:
            raise ValueError(f"No existe la columna requerida para aplicar fecha de cierre: {columna}")

    base_cierre = base.copy()

    base_cierre["Fecha_Programada"] = pd.to_datetime(
        base_cierre["Fecha_Programada"],
        errors="coerce",
        dayfirst=True
    )

    base_cierre["FechaPago"] = pd.to_datetime(
        base_cierre["FechaPago"],
        errors="coerce",
        dayfirst=True
    )

    base_cierre["Monto_Cuota_Pagado"] = pd.to_numeric(
        base_cierre["Monto_Cuota_Pagado"],
        errors="coerce"
    ).fillna(0)

    base_cierre["SaldoPorPagarCuota"] = pd.to_numeric(
        base_cierre["SaldoPorPagarCuota"],
        errors="coerce"
    ).fillna(0)

    if "Monto_Cuota" in base_cierre.columns:
        base_cierre["Monto_Cuota"] = pd.to_numeric(
            base_cierre["Monto_Cuota"],
            errors="coerce"
        ).fillna(0)

    condicion_pago_posterior_cierre = (
        (base_cierre["FechaPago"].notna())
        & (base_cierre["FechaPago"] > fecha_cierre)
        & (base_cierre["Monto_Cuota_Pagado"] != 0)
    )

    cantidad_pagos_posteriores = int(condicion_pago_posterior_cierre.sum())

    if cantidad_pagos_posteriores > 0:
        if "Monto_Cuota" in base_cierre.columns:
            saldo_restaurado = base_cierre["Monto_Cuota"].copy().fillna(0)
            # Si la fila tiene cuota programada cero, no se crea deuda artificial.
            saldo_restaurado = saldo_restaurado.mask(saldo_restaurado < 0, 0)
        else:
            saldo_restaurado = base_cierre["SaldoPorPagarCuota"].copy().fillna(0)

        base_cierre.loc[condicion_pago_posterior_cierre, "Monto_Cuota_Pagado"] = 0
        base_cierre.loc[condicion_pago_posterior_cierre, "SaldoPorPagarCuota"] = saldo_restaurado.loc[condicion_pago_posterior_cierre]

    base_cierre, cantidad_vencidos, cantidad_pendientes = reclasificar_estado_por_fecha_cierre(
        base_cierre,
        fecha_cierre
    )

    print("\n====================================")
    print("AJUSTE POR FECHA DE CIERRE")
    print("====================================")
    print(f"Fecha de cierre: {fecha_cierre.strftime('%d/%m/%Y')}")
    print(f"Pagos posteriores al cierre revertidos en Base_Ajustada: {cantidad_pagos_posteriores}")
    print(f"Filas con saldo reclasificadas como VENCIDO: {cantidad_vencidos}")
    print(f"Filas con saldo reclasificadas como PENDIENTE: {cantidad_pendientes}")

    return base_cierre

def aplicar_ajuste_saldos_negativos(base):
    """
    Reestructura saldos negativos para el reporte SIN cambiar la recaudación.
    """
    columnas_necesarias = [
        "Proyecto",
        "Fecha_Programada",
        "SaldoPorPagarCuota",
        "Monto_Cuota_Pagado",
    ]

    for columna in columnas_necesarias:
        if columna not in base.columns:
            raise ValueError(f"No existe la columna requerida para el ajuste: {columna}")

    columna_cliente = obtener_columna_cliente_ajuste(base)
    columna_departamento = obtener_columna_departamento_ajuste(base)

    base_ajustada = base.copy()

    # Asegura tipos correctos para el cálculo.
    base_ajustada["Fecha_Programada"] = pd.to_datetime(
        base_ajustada["Fecha_Programada"],
        errors="coerce",
        dayfirst=True
    )

    base_ajustada["SaldoPorPagarCuota"] = pd.to_numeric(
        base_ajustada["SaldoPorPagarCuota"],
        errors="coerce"
    ).fillna(0)

    base_ajustada["_Orden_Original_Ajuste"] = range(len(base_ajustada))

    columnas_grupo = [
        "Proyecto",
        columna_cliente,
        columna_departamento,
    ]

    total_negativos = int((base_ajustada["SaldoPorPagarCuota"] < 0).sum())
    negativos_aplicados = 0
    monto_reestructurado = 0.0

    # Solo se ajustan filas que tienen la llave completa.
    llave_completa = base_ajustada[columnas_grupo].notna().all(axis=1)
    for columna in columnas_grupo:
        llave_completa = llave_completa & (base_ajustada[columna].astype(str).str.strip() != "")

    indices_validos = base_ajustada[llave_completa].index

    grupos = base_ajustada.loc[indices_validos].groupby(columnas_grupo, dropna=False, sort=False)

    for _, grupo in grupos:
        grupo_ordenado = grupo.sort_values(
            by=["Fecha_Programada", "_Orden_Original_Ajuste"],
            ascending=[True, True]
        )

        indices_grupo = list(grupo_ordenado.index)

        for posicion, idx_negativo in enumerate(indices_grupo):
            saldo_negativo = float(base_ajustada.at[idx_negativo, "SaldoPorPagarCuota"])

            if saldo_negativo >= 0:
                continue

            credito_pendiente = abs(saldo_negativo)
            credito_inicial = credito_pendiente

            # Buscar cuotas posteriores del mismo cliente/departamento/proyecto.
            for idx_destino in indices_grupo[posicion + 1:]:
                if credito_pendiente <= 0:
                    break

                saldo_destino = float(base_ajustada.at[idx_destino, "SaldoPorPagarCuota"])

                if saldo_destino <= 0:
                    continue

                monto_a_aplicar = min(credito_pendiente, saldo_destino)

                base_ajustada.at[idx_destino, "SaldoPorPagarCuota"] = saldo_destino - monto_a_aplicar
                credito_pendiente -= monto_a_aplicar
                monto_reestructurado += monto_a_aplicar

            # Si el crédito se aplicó completo, la fila negativa queda en cero.
            # Si no se encontró saldo futuro suficiente, se conserva el remanente negativo
            # para no perder el cuadre total de la base.
            if credito_pendiente <= 0:
                base_ajustada.at[idx_negativo, "SaldoPorPagarCuota"] = 0
                negativos_aplicados += 1
            else:
                base_ajustada.at[idx_negativo, "SaldoPorPagarCuota"] = -credito_pendiente
                if credito_pendiente < credito_inicial:
                    negativos_aplicados += 1

    base_ajustada = base_ajustada.drop(columns=["_Orden_Original_Ajuste"])

    print("\n====================================")
    print("AJUSTE DE SALDOS NEGATIVOS")
    print("====================================")
    print(f"Saldos negativos encontrados: {total_negativos}")
    print(f"Filas negativas reestructuradas total/parcial: {negativos_aplicados}")
    print(f"Monto aplicado a cuotas futuras: {monto_reestructurado:,.2f}")
    print("Regla: mismo proyecto, mismo cliente, mismo departamento y orden por Fecha_Programada.")

    return base_ajustada



def aplicar_estado_parcial_cancelado(base):
    """
    Marca cuotas parcialmente pagadas en Base_Ajustada.
    """
    fecha_cierre = obtener_fecha_cierre_ajuste()

    base_parcial = base.copy()

    columnas_necesarias = [
        "Estado",
        "Fecha_Programada",
        "Monto_Cuota_Pagado",
        "SaldoPorPagarCuota",
    ]

    for columna in columnas_necesarias:
        if columna not in base_parcial.columns:
            raise ValueError(f"No existe la columna requerida para marcar parciales: {columna}")

    base_parcial["Fecha_Programada"] = pd.to_datetime(
        base_parcial["Fecha_Programada"],
        errors="coerce",
        dayfirst=True
    )

    base_parcial["Monto_Cuota_Pagado"] = pd.to_numeric(
        base_parcial["Monto_Cuota_Pagado"],
        errors="coerce"
    ).fillna(0)

    base_parcial["SaldoPorPagarCuota"] = pd.to_numeric(
        base_parcial["SaldoPorPagarCuota"],
        errors="coerce"
    ).fillna(0)

    condicion_parcial = (
        (base_parcial["Monto_Cuota_Pagado"] > 0)
        & (base_parcial["SaldoPorPagarCuota"] > 0)
    )

    if fecha_cierre is not None:
        condicion_vencido = (
            condicion_parcial
            & base_parcial["Fecha_Programada"].notna()
            & (base_parcial["Fecha_Programada"] <= fecha_cierre)
        )

        condicion_pendiente = (
            condicion_parcial
            & base_parcial["Fecha_Programada"].notna()
            & (base_parcial["Fecha_Programada"] > fecha_cierre)
        )

    else:
        estado_actual = base_parcial["Estado"].astype(str).apply(normalizar_texto)
        condicion_vencido = condicion_parcial & estado_actual.str.contains("VENCIDO", na=False)
        condicion_pendiente = condicion_parcial & estado_actual.str.contains("PENDIENTE", na=False)

    base_parcial.loc[condicion_vencido, "Estado"] = "VENCIDO - PARC. CANCELADO"
    base_parcial.loc[condicion_pendiente, "Estado"] = "PENDIENTE - PARC. CANCELADO"

    cantidad_parcial = int((condicion_vencido | condicion_pendiente).sum())

    print("\n====================================")
    print("ESTADO PARCIAL CANCELADO")
    print("====================================")
    print(f"Cuotas parcialmente canceladas marcadas: {cantidad_parcial}")
    print("Regla: Monto_Cuota_Pagado > 0 y SaldoPorPagarCuota > 0.")

    return base_parcial


def filtrar_ventas_hasta_fecha_cierre(base):
    """
    Filtra Base_Ajustada para el cierre mensual.
    """
    fecha_cierre = obtener_fecha_cierre_ajuste()

    if fecha_cierre is None:
        print("ℹ️ No se filtró FechaOperacionComercial porque FECHA_CORTE_REPORTE está vacío.")
        return base.copy()

    if "FechaOperacionComercial" not in base.columns:
        raise ValueError(
            "❌ No se puede filtrar la base ajustada porque no existe la columna "
            "FechaOperacionComercial."
        )

    base_filtrada = base.copy()

    fecha_operacion = pd.to_datetime(
        base_filtrada["FechaOperacionComercial"],
        errors="coerce",
        dayfirst=True
    )

    condicion_mantener = (
        fecha_operacion.notna()
        & (fecha_operacion <= fecha_cierre)
    )

    filas_antes = len(base_filtrada)
    filas_eliminadas = int((~condicion_mantener).sum())
    filas_posteriores = int((fecha_operacion.notna() & (fecha_operacion > fecha_cierre)).sum())
    filas_sin_fecha = int(fecha_operacion.isna().sum())

    base_filtrada = base_filtrada.loc[condicion_mantener].copy()

    print("\n====================================")
    print("FILTRO DE VENTAS POR FECHA DE CIERRE")
    print("====================================")
    print(f"Fecha de cierre: {fecha_cierre.strftime('%d/%m/%Y')}")

    return base_filtrada

def aplicar_formatos_base_ajustada(ruta_excel):
    from openpyxl import load_workbook

    wb = load_workbook(ruta_excel)

    for nombre_hoja in ["Base_Original", "Base_Ajustada"]:
        if nombre_hoja not in wb.sheetnames:
            continue

        ws = wb[nombre_hoja]

        encabezados = {}
        for celda in ws[1]:
            encabezados[celda.value] = celda.column

        for columna in COLUMNAS_FECHA:
            if columna in encabezados:
                col_idx = encabezados[columna]
                for fila in range(2, ws.max_row + 1):
                    ws.cell(row=fila, column=col_idx).number_format = "dd/mm/yyyy"

        for columna in encabezados:
            if es_columna_numero(columna):
                col_idx = encabezados[columna]
                for fila in range(2, ws.max_row + 1):
                    ws.cell(row=fila, column=col_idx).number_format = "#,##0.00"

        ws.freeze_panes = "A2"

    wb.save(ruta_excel)


def generar_base_cobranzas_ajustada(base_consolidada):
    """
    Genera un archivo adicional en Flujo/Output:
        Base_Cobranzas_Ajustado.xlsx

    Contiene:
        - Base_Original:
          Base consolidada tal como queda luego de la descarga/consolidación.
          NO se filtra por FechaOperacionComercial.
          NO se aplica fecha de cierre.
          NO se reclasifican estados.
          NO se reestructuran saldos negativos.
          Solo conserva los formatos de fecha/número al exportar.

        - Base_Ajustada:
          Base adaptada para el cierre mensual del reporte.
          Aquí SÍ se aplican los ajustes:
          1. Fecha de cierre.
          2. Reclasificación VENCIDO/PENDIENTE.
          3. Reestructuración de saldos negativos.
          4. Estado parcial cancelado.
          5. Filtro de FechaOperacionComercial <= FECHA_CORTE_REPORTE.
    """

    # IMPORTANTE:
    # La base original se separa desde el inicio y no se vuelve a tocar.
    # Los ajustes se hacen únicamente sobre base_ajustada.
    base_original = base_consolidada.copy()
    base_ajustada = base_consolidada.copy()

    # 1) Fotografía al cierre del reporte.
    #    Ejemplo: si el cierre es junio, los pagos posteriores al cierre
    #    ya no cuentan como cobrados dentro de Base_Ajustada.
    base_ajustada = aplicar_fecha_cierre_base_ajustada(base_ajustada)

    # 2) Reestructura saldos negativos contra cuotas futuras del mismo
    #    proyecto, cliente y departamento, sin cambiar Monto_Cuota_Pagado.
    base_ajustada = aplicar_ajuste_saldos_negativos(base_ajustada)

    # 3) Reclasifica el estado por fecha de cierre usando el saldo ya ajustado.
    fecha_cierre = obtener_fecha_cierre_ajuste()
    base_ajustada, _, _ = reclasificar_estado_por_fecha_cierre(base_ajustada, fecha_cierre)

    # 4) Marca cuotas parcialmente canceladas sin modificar montos.
    #    Mantiene VENCIDO/PENDIENTE dentro del Estado para que los reportes
    #    sigan funcionando con los filtros existentes.
    base_ajustada = aplicar_estado_parcial_cancelado(base_ajustada)

    # 5) Filtro final SOLO en Base_Ajustada.
    #    Base_Original NO se filtra.
    base_ajustada = filtrar_ventas_hasta_fecha_cierre(base_ajustada)

    ruta_base_ajustada = SALIDA_LISTA / NOMBRE_BASE_AJUSTADA

    with pd.ExcelWriter(ruta_base_ajustada, engine="openpyxl") as writer:
        base_original.to_excel(
            writer,
            sheet_name="Base_Original",
            index=False
        )

        base_ajustada.to_excel(
            writer,
            sheet_name="Base_Ajustada",
            index=False
        )

    aplicar_formatos_base_ajustada(ruta_base_ajustada)

    print("\n✅ Base ajustada generada correctamente:")
    print(ruta_base_ajustada.resolve())

    return ruta_base_ajustada

# =========================================================
# 8. DESCARGA POR RANGO
# =========================================================

def descargar_rango(driver, anio, fecha_inicio, fecha_fin):
    ultimo_error = None

    for intento in range(1, INTENTOS_DESCARGA + 1):
        print(f"Intento de descarga {intento} de {INTENTOS_DESCARGA}")

        try:
            archivos_antes = obtener_archivos_downloads()

            seleccionar_formato(driver, FORMATO_EVOLTA)

            escribir_fecha(driver, "Fecha de inicio", fecha_inicio)
            escribir_fecha(driver, "Fecha de fin", fecha_fin)

            tiempo_inicio_descarga = time.time()

            click_exportar(driver)

            archivo_descargado_downloads = esperar_nueva_descarga_downloads(
                archivos_antes=archivos_antes,
                tiempo_inicio=tiempo_inicio_descarga,
                timeout=TIEMPO_ESPERA_DESCARGA
            )

            return archivo_descargado_downloads

        except Exception as e:
            ultimo_error = e
            print(f"No se detectó descarga en el intento {intento}.")
            print(e)

            if intento < INTENTOS_DESCARGA:
                print("Recargando página de reporte para intentar nuevamente...")
                ir_a_reporte_cobranzas(driver)
                time.sleep(2)

    raise ultimo_error


# =========================================================
# 9. PROGRAMA PRINCIPAL
# =========================================================

def main():
    print("====================================")
    print("AUTOMATIZACIÓN COBRANZAS EVOLTA")
    print("====================================")

    if not DOWNLOADS_DIR.exists():
        raise FileNotFoundError(f"No existe la carpeta de Descargas: {DOWNLOADS_DIR}")

    ENTRADA_COBRANZAS.mkdir(parents=True, exist_ok=True)

    rangos = generar_rangos_por_anio(FECHA_INICIO_REPORTE, FECHA_FIN_REPORTE)

    print("\nRangos que se van a descargar:")

    for r in rangos:
        print(f"{r['anio']}: {r['fecha_inicio']} al {r['fecha_fin']}")

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")

    prefs = {
        "download.default_directory": str(DOWNLOADS_DIR.resolve()),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_settings.popups": 0,
        "profile.default_content_setting_values.automatic_downloads": 1
    }

    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)
    configurar_descargas_chrome(driver)

    rangos_con_error = []

    try:
        if LOGIN_AUTOMATICO:
            iniciar_sesion_evolta(driver)
        else:
            driver.get(LOGIN_URL)

            input("\nInicia sesión en Evolta y luego presiona ENTER aquí... ")

            ir_a_reporte_cobranzas(driver)

        for r in rangos:
            anio = r["anio"]
            fecha_inicio = r["fecha_inicio"]
            fecha_fin = r["fecha_fin"]

            print("\n====================================")
            print(f"PROCESANDO AÑO {anio}")
            print(f"Fecha inicio: {fecha_inicio}")
            print(f"Fecha fin: {fecha_fin}")
            print("====================================")

            try:
                limpiar_archivos_previos_anio(anio)

                archivo_descargado_downloads = descargar_rango(
                    driver=driver,
                    anio=anio,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin
                )

                archivo_renombrado_downloads = renombrar_en_downloads(
                    archivo_descargado=archivo_descargado_downloads,
                    anio=anio
                )

                archivo_final = copiar_renombrado_a_cobranzas(
                    archivo_renombrado=archivo_renombrado_downloads
                )

                if CONVERTIR_CSV_A_EXCEL and archivo_final.suffix.lower() == ".csv":
                    convertir_csv_a_excel_seguro(archivo_final)

                print(f"Año {anio} descargado correctamente.")

                time.sleep(2)

            except Exception as e:
                print(f"No se pudo descargar el año {anio}.")
                print("Motivo:")
                print(e)
                print("Se continuará con el siguiente año.")

                rangos_con_error.append({
                    "anio": anio,
                    "fecha_inicio": fecha_inicio,
                    "fecha_fin": fecha_fin,
                    "error": str(e)
                })

                try:
                    ir_a_reporte_cobranzas(driver)
                except Exception:
                    pass

                continue

        print("\n====================================")
        print("PROCESO DE DESCARGA TERMINADO")
        print("====================================")

        if rangos_con_error:
            print("\nRANGOS NO DESCARGADOS:")

            for error in rangos_con_error:
                print(
                    f"{error['anio']}: "
                    f"{error['fecha_inicio']} al {error['fecha_fin']} - "
                    f"{error['error']}"
                )

        else:
            print("Todos los rangos fueron descargados correctamente.")

        print("\nArchivos guardados en:")
        print(ENTRADA_COBRANZAS.resolve())

        ruta_base_consolidada = generar_base_cobranzas_consolidada()

        print("\nBase consolidada guardada en:")
        print(ruta_base_consolidada.resolve())

        base_consolidada = pd.read_excel(
            ruta_base_consolidada,
            sheet_name=NOMBRE_HOJA_BASE,
            engine="openpyxl"
        )
        base_consolidada = convertir_tipos_base(base_consolidada)

        ruta_base_ajustada = generar_base_cobranzas_ajustada(base_consolidada)

        print("\nBase ajustada guardada en:")
        print(ruta_base_ajustada.resolve())

    except Exception as e:
        print("\nOcurrió un error general:")
        print(e)

    finally:
        driver.quit()


if __name__ == "__main__":
    main()