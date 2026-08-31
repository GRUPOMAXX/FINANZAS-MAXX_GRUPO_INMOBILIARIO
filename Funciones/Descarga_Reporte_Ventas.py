# Programa para descargar el reporte de ventas con fecha de corte
# Creado por Eduardo Miguel Huamani Acosta                     31/08/26

from pathlib import Path
import os
import sys
import time
import shutil
import pandas as pd
import warnings

warnings.filterwarnings("ignore")

from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# =====================================================
# CONFIGURACION
# =====================================================

BASE_DIR = Path(__file__).resolve().parent.parent

if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from Conexiones.connection import FECHA_CORTE_REPORTE


LOGIN_URL = "https://v4.evolta.pe/Login/Acceso/Index"

REPORTE_STOCK_URL = ("https://v4.evolta.pe/Reportes/RepCargaStock/Index")


DOWNLOADS_DIR = Path(r"C:\Users\ehuamani\Downloads")


ENTRADA_VENTAS = (BASE_DIR /"Flujo" /"Input" /"Reporte de Ventas")


SALIDA_VENTAS = (BASE_DIR /"Flujo" /"Output")


ENTRADA_VENTAS.mkdir(parents=True, exist_ok=True)
SALIDA_VENTAS.mkdir(parents=True, exist_ok=True)


ARCHIVO_VENTAS = (SALIDA_VENTAS /"Reporte_Ventas.xlsx")


load_dotenv(BASE_DIR / ".env")

USUARIO_EVOLTA = os.getenv("USUARIO_EVOLTA")
CLAVE_EVOLTA = os.getenv("CLAVE_EVOLTA")


# =====================================================
# DRIVER
# =====================================================

def crear_driver_chrome():

    options = Options()

    options.add_argument(
        "--start-maximized"
    )

    options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": str(DOWNLOADS_DIR),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True
        }
    )

    return webdriver.Chrome(options=options)


# =====================================================
# UTILIDADES
# =====================================================

def click_seguro(driver, elemento):

    try:
        elemento.click()

    except:

        try:
            ActionChains(driver).move_to_element(elemento).click().perform()

        except:
            driver.execute_script(
                "arguments[0].click();",
                elemento
            )


# =====================================================
# LOGIN
# =====================================================

def iniciar_sesion_evolta(driver):

    print("💻 Ingresando a Evolta...")

    driver.get(LOGIN_URL)

    wait = WebDriverWait(driver, 40)

    usuario = wait.until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "//input[@type='email'] | //input[not(@type='password') and not(@type='hidden')][1]"
            )
        )
    )

    clave = wait.until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "//input[@type='password']"
            )
        )
    )

    usuario.send_keys(USUARIO_EVOLTA)
    clave.send_keys(CLAVE_EVOLTA)

    boton = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                "//button[contains(.,'Acceder')] | //button[contains(.,'Ingresar')] | //input[@type='submit']"
            )
        )
    )

    click_seguro(driver, boton)

    time.sleep(5)

    print("✅ Inicio de sesion correcto")


# =====================================================
# REPORTE STOCK
# =====================================================

def ir_a_stock(driver):

    print("🔎 Ingresando a Inmuebles...")

    driver.get(REPORTE_STOCK_URL)

    time.sleep(8)


def seleccionar_excel(driver):

    wait = WebDriverWait(driver, 40)

    excel = wait.until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "//*[normalize-space()='Excel']/preceding::input[@type='radio'][1]"
            )
        )
    )

    driver.execute_script(
        "arguments[0].click();",
        excel
    )


def click_exportar(driver):

    wait = WebDriverWait(driver, 40)

    boton = wait.until(
        EC.element_to_be_clickable(
            (
                By.XPATH,
                "//button[contains(.,'Exportar')] | //input[@value='Exportar'] | //a[contains(.,'Exportar')]"
            )
        )
    )

    click_seguro(driver, boton)

    time.sleep(10)


# =====================================================
# DESCARGA (LOGICA IGUAL A COBRANZAS)
# =====================================================

def obtener_ultimo_excel_descargado():

    archivos = list(
        DOWNLOADS_DIR.glob("*.xlsx")
    )

    if not archivos:
        raise Exception(
            "❌ No se encontró archivo Excel descargado"
        )

    return max(
        archivos,
        key=lambda x: x.stat().st_mtime
    )


def renombrar_en_downloads_y_copiar_input(archivo):

    archivo_renombrado = (
        DOWNLOADS_DIR /
        "Reporte_Ventas.xlsx"
    )

    if archivo_renombrado.exists():
        archivo_renombrado.unlink()

    archivo.rename(
        archivo_renombrado
    )


    destino = (
        ENTRADA_VENTAS /
        "Reporte_Ventas.xlsx"
    )

    if destino.exists():
        destino.unlink()


    shutil.copy2(
        str(archivo_renombrado),
        str(destino)
    )


    return destino


# =====================================================
# HISTORICO
# =====================================================

def actualizar_historico(archivo):

    print("⌛ Procesando reporte...")

    nuevo = pd.read_excel(
        archivo
    )

    nuevo["FechaCorte"] = pd.to_datetime(
        FECHA_CORTE_REPORTE,
        dayfirst=True
    )


    if ARCHIVO_VENTAS.exists():

        historico = pd.read_excel(
            ARCHIVO_VENTAS
        )


        historico["FechaCorte"] = pd.to_datetime(
            historico["FechaCorte"]
        )


        fecha = pd.to_datetime(
            FECHA_CORTE_REPORTE,
            dayfirst=True
        )


        historico = historico[
            historico["FechaCorte"] != fecha
        ]


        consolidado = pd.concat(
            [
                historico,
                nuevo
            ],
            ignore_index=True
        )


    else:

        consolidado = nuevo


    consolidado.to_excel(
        ARCHIVO_VENTAS,
        index=False
    )

    from openpyxl import load_workbook

    wb = load_workbook(ARCHIVO_VENTAS)
    ws = wb.active

    columna_fecha = None

    for celda in ws[1]:
        if celda.value == "FechaCorte":
            columna_fecha = celda.column
            break

    if columna_fecha:
        for fila in range(2, ws.max_row + 1):
            ws.cell(fila, columna_fecha).number_format = "dd/mm/yyyy"

    wb.save(ARCHIVO_VENTAS)


    print("✅ Archivo generado:")
    print(ARCHIVO_VENTAS)


# =====================================================
# MAIN
# =====================================================

def main():

    driver = crear_driver_chrome()

    try:

        iniciar_sesion_evolta(driver)

        ir_a_stock(driver)

        seleccionar_excel(driver)

        click_exportar(driver)

        time.sleep(10)

        archivo = obtener_ultimo_excel_descargado()


    finally:

        driver.quit()


    archivo_input = renombrar_en_downloads_y_copiar_input(
        archivo
    )


    actualizar_historico(
        archivo_input
    )


    print("📁 Proceso terminado correctamente")


if __name__ == "__main__":
    main()
