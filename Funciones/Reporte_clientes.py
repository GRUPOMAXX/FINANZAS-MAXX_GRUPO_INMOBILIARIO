# Programa para reporte de cobranzas por cliente
# Creado por Eduardo Miguel Huamani Acosta                             16/07/26

# Esta automatización trata de elaborar un reporte por cliente y por proyecto. Al momento de elaborar el reporte se tendría que
# quitar los duplicados, sumar los montos, lo cual es complicado formular en excel. 

# Este código permite extraer información de las cobranzas y elabora un reporte de clientes morosos, ahorrandose el 
# trabajo de formular, quitar duplicados. Este reporte de lo que se hacía en 2 - 3 horas, 
# con el codigo se podrá tener el reporte completo en 2 - 3 minutos.

# =============================================================================================================================

#LIBRERIAS

import sys
import re
import unicodedata
from pathlib import Path
from copy import copy
from statistics import mean

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.formatting import ConditionalFormattingList


BASE_DIR = Path(__file__).resolve().parent.parent

if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from Conexiones.connection import (  # noqa: E402
    FECHA_CORTE_REPORTE,
    PROYECTO,
)

try:
    from Conexiones.connection import PROYECTOS_VIGENTES_COBRANZAS  # noqa: E402
except ImportError:
    PROYECTOS_VIGENTES_COBRANZAS = []

import Descarga_Cobranzas as descarga  # noqa: E402


# =========================================================
# 1. VARIABLES DEL REPORTE
# =========================================================

EJECUTAR_DESCARGA = True

SALIDA_LISTA = BASE_DIR / "Flujo" / "Output"
CARPETA_PLANTILLAS = BASE_DIR / "Flujo" / "Input" / "Plantillas"

NOMBRE_BASE_AJUSTADA = "Base_Cobranzas_Ajustado.xlsx"
NOMBRE_HOJA_BASE_AJUSTADA = "Base_Ajustada"

NOMBRE_PLANTILLA_GERENCIA = "Plantilla_Cobranza_por_cliente_para_gerencia.xlsx"
NOMBRE_REPORTE_GERENCIA = "Reporte_para_Gerencia.xlsx"

HOJA_REPORTE_GENERAL = "Reporte_General"
HOJA_REPORTE_CLIENTE = "Reporte_Cliente"
HOJA_REPORTE_PROYECTO = "Reporte_Proyecto"

FORMATO_MONTO = '#,##0;-#,##0;-'
FORMATO_PORCENTAJE = '0.0%'
FORMATO_DIAS = '#,##0;-#,##0;-'

# Excel maneja alto de fila en puntos. 21.6 puntos ≈ 28.8 px.
ALTO_FILA_CLIENTE_PUNTOS = 28.8

SALIDA_LISTA.mkdir(parents=True, exist_ok=True)

# Rangos donde se aplicarán barras de datos al final con Excel nativo.
# Se usa Excel COM para evitar archivos con aviso de recuperación.
RANGOS_BARRAS_DEUDA = []


# Diccionario oficial de proyectos importado desde Conexiones/connection.py.
# No se usa PROYECTO_DEFAULT ni fallback local.

ORDEN_PROYECTOS = list(dict.fromkeys([
    str(nombre_comercial).strip()
    for nombre_comercial in PROYECTO.values()
    if str(nombre_comercial).strip() != ""
]))

COLUMNAS_FECHA = [
    "Fecha_Programada",
    "FechaPago",
]

COLUMNAS_NUMERO_FIJAS = [
    "Monto_Cuota",
    "Monto_Cuota_Pagado",
    "SaldoPorPagarCuota",
]

COLUMNAS_NUMERO_CON_INDICE = [
    "PrecioLista",
    "PrecioVenta",
]


# =========================================================
# 2. FUNCIONES GENERALES
# =========================================================

# Ejecuta el script Descarga_Cobranzas
def ejecutar_descarga_cobranzas():
    """
    Ejecuta Descarga_Cobranzas.py.
    """
    descarga.main()

# Limpia textos para comparar sin errores
def normalizar_texto(texto):
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )
    return texto

# Convierte el nombre largo del proyecto al nombre comercial.
def obtener_nombre_proyecto(proyecto):
    proyecto_normalizado = normalizar_texto(proyecto)

    for nombre_original, nombre_formato in PROYECTO.items():
        if normalizar_texto(nombre_original) == proyecto_normalizado:
            return nombre_formato

    for nombre_original, nombre_formato in PROYECTO.items():
        if normalizar_texto(nombre_original) in proyecto_normalizado:
            return nombre_formato

    return str(proyecto).strip()

# Convierte textos a números
def convertir_columna_numero(serie):
    serie = serie.astype(str).str.strip()
    serie = serie.str.replace("S/", "", regex=False)
    serie = serie.str.replace("US$", "", regex=False)
    serie = serie.str.replace("$", "", regex=False)
    serie = serie.str.replace(" ", "", regex=False)
    serie = serie.str.replace(",", "", regex=False)
    serie = serie.replace("", pd.NA)
    return pd.to_numeric(serie, errors="coerce")

# Identifica qué columnas tienen que ser numericas
def es_columna_numero(columna):
    if columna in COLUMNAS_NUMERO_FIJAS:
        return True

    for campo in COLUMNAS_NUMERO_CON_INDICE:
        if re.match(rf"^{campo}_(\d+)$", str(columna)):
            return True

    return False

# Convierte fechas y montos de la base para que se puedan calcular correctamente.
def convertir_tipos_base(base):
    base = base.copy()

    for columna in COLUMNAS_FECHA:
        if columna in base.columns:
            base[columna] = pd.to_datetime(
                base[columna],
                errors="coerce",
                dayfirst=True,
            )

    for columna in base.columns:
        if es_columna_numero(columna):
            base[columna] = convertir_columna_numero(base[columna])

    return base

# Lee la fecha de corte y lo cambia a fecha real
def obtener_fecha_corte():
    fecha_corte = pd.to_datetime(
        FECHA_CORTE_REPORTE,
        errors="coerce",
        dayfirst=True,
    )

    if pd.isna(fecha_corte):
        raise ValueError(
            "FECHA_CORTE_REPORTE no tiene formato válido. "
            "Usa un formato como '30/06/2026'."
        )

    return fecha_corte

# Busca qué columna se usará para identificar al cliente
def obtener_columna_cliente(base):
    candidatos = [
        "NroDocumento",
        "Nro_Documento",
        "Documento",
        "NumeroDocumento",
        "NumDocumento",
        "DNI",
        "Nombres_Titular",
        "Cliente",
    ]

    for col in candidatos:
        if col in base.columns:
            return col

    raise ValueError(
        "No se encontró columna para identificar al cliente. "
        "Debe existir NroDocumento, Documento, Nombres_Titular o Cliente."
    )

# Devuelve el nombre visible del cliente.
def obtener_nombre_cliente(row):
    candidatos = [
        "Nombres_Titular",
        "Cliente",
        "NombreCliente",
        "Nombre_Titular",
    ]

    for col in candidatos:
        if col in row.index:
            valor = row.get(col)
            if pd.notna(valor) and str(valor).strip() != "":
                return str(valor).strip()

    return "Sin cliente"

# Detecta cuántas columnas de inmueble existen (TipoInmueble)
def obtener_indices_inmueble(base):
    indices = set()

    for columna in base.columns:
        coincidencia = re.match(r"^TipoInmueble_(\d+)$", str(columna))
        if coincidencia:
            indices.add(int(coincidencia.group(1)))

    return sorted(indices)

# Detecta cuántas columnas de inmueble existen
def limpiar_tipo_inmueble(tipo):
    tipo_norm = normalizar_texto(tipo)

    if "DEPART" in tipo_norm:
        return "Dpto"
    if "ESTACION" in tipo_norm:
        return "Estc"
    if "DEPOS" in tipo_norm:
        return "Depósito"
    if "LOCAL" in tipo_norm or "COMERCIAL" in tipo_norm:
        return "Local"

    return str(tipo).strip()

# Obtiene las unidades de una fila
def obtener_unidades_fila(row, indices):
    unidades = []

    for idx in indices:
        tipo = row.get(f"TipoInmueble_{idx}", "")
        nro = row.get(f"NroInmueble_{idx}", "")

        if pd.isna(tipo) or str(tipo).strip() == "":
            continue

        if pd.isna(nro) or str(nro).strip() == "":
            continue

        tipo_limpio = limpiar_tipo_inmueble(tipo)
        nro_limpio = str(nro).strip()

        precio = row.get(f"PrecioVenta_{idx}", 0)
        precio = pd.to_numeric(precio, errors="coerce")
        precio = 0 if pd.isna(precio) else float(precio)

        clave = f"{normalizar_texto(tipo_limpio)}|{normalizar_texto(nro_limpio)}"

        unidades.append({
            "clave": clave,
            "label": f"{tipo_limpio} {nro_limpio}",
            "precio": precio,
        })

    if not unidades:
        unidades.append({
            "clave": "SIN_UNIDAD",
            "label": "Sin unidad",
            "precio": 0.0,
        })

    return unidades

# Crea una clave interna para agrupar una unidad o conjunto de unidades.
def obtener_clave_unidad(unidades):
    claves = [u["clave"] for u in unidades]
    return " | ".join(claves)

# Crea el texto visible de la unidad. Ej: Dpto-205
def obtener_label_unidad(unidades):
    labels = []
    vistos = set()

    for unidad in unidades:
        if unidad["clave"] in vistos:
            continue

        labels.append(unidad["label"])
        vistos.add(unidad["clave"])

    return ", ".join(labels)

# Obtiene el tipo de financiamiento del cliente
def obtener_financiamiento(row):
    candidatos = [
        "TipoFinanciamiento",
        "Tipo_Financiamiento",
        "Financiamiento",
    ]

    for col in candidatos:
        if col in row.index:
            valor = row.get(col)
            if pd.notna(valor) and str(valor).strip() != "":
                return str(valor).strip()

    return ""

# Filtra la base para dejar solo los proyectos vigentes
def filtrar_proyectos_vigentes(df):
    if not PROYECTOS_VIGENTES_COBRANZAS:
        return df

    proyectos_permitidos = {
        normalizar_texto(obtener_nombre_proyecto(proyecto))
        for proyecto in PROYECTOS_VIGENTES_COBRANZAS
    }

    return df[
        df["Proyecto"].astype(str).apply(obtener_nombre_proyecto).apply(normalizar_texto).isin(proyectos_permitidos)
    ].copy()

# Busca qué columna representa la deuda: 
def obtener_columna_deuda_resumen(resumen_cliente):
    posibles_columnas = [
        "Vencido",
        "Deuda",
        "Monto Deuda",
        "Monto_Deuda",
    ]

    for columna in posibles_columnas:
        if columna in resumen_cliente.columns:
            return columna

    raise ValueError(
        "❌ No se encontró columna de deuda en resumen_cliente. "
        "Revisa si se llama Vencido, Deuda o Monto Deuda."
    )

# Une textos
def unir_texto_con_y(elementos):
    elementos = [str(x).strip() for x in elementos if str(x).strip() != ""]

    if not elementos:
        return ""

    if len(elementos) == 1:
        return elementos[0]

    if len(elementos) == 2:
        return f"{elementos[0]} y {elementos[1]}"

    return ", ".join(elementos[:-1]) + f" y {elementos[-1]}"

# Detecta clientes morosos que aparecen con deuda en más de un proyecto.
def detectar_clientes_morosos_multiproyecto(resumen_cliente):
    """
    Detecta clientes morosos que aparecen con deuda vencida en más de un proyecto.
    """

    if resumen_cliente.empty:
        return []

    columna_deuda = obtener_columna_deuda_resumen(resumen_cliente)

    base = resumen_cliente.copy()

    if "Cliente_ID" not in base.columns:
        base["Cliente_ID"] = base["Cliente"].astype(str).str.strip().str.upper()

    base[columna_deuda] = pd.to_numeric(
        base[columna_deuda],
        errors="coerce",
    ).fillna(0)

    deudores = base[base[columna_deuda] > 0].copy()

    if deudores.empty:
        return []

    alertas = []

    for cliente_id, temp in deudores.groupby("Cliente_ID"):
        proyectos = sorted(temp["Proyecto"].dropna().astype(str).str.strip().unique())

        if len(proyectos) <= 1:
            continue

        cliente = str(temp["Cliente"].dropna().iloc[0]).strip()

        alertas.append({
            "Cliente": cliente,
            "Cliente_ID": cliente_id,
            "Proyectos": proyectos,
        })

    return alertas

# Construye el texto en donde se menciona que un cliente debe mas de un proyecto
def construir_texto_nota_multiproyecto(resumen_cliente):
    alertas = detectar_clientes_morosos_multiproyecto(resumen_cliente)

    if not alertas:
        return ""

    textos = []

    for alerta in alertas:
        cliente = alerta["Cliente"]
        proyectos_texto = unir_texto_con_y(alerta["Proyectos"])

        textos.append(
            f"{cliente} es cliente de {proyectos_texto}."
        )

    return " ".join(textos)

# Escribe la nota debajo del TOTAL en la hoja Reporte_Proyecto
def agregar_nota_multiproyecto_reporte_proyecto(ws, resumen_cliente, fila_total):
    """
    Agrega una nota debajo de la fila TOTAL en Reporte_Proyecto.

    La nota aparece solo si existe un cliente moroso en más de un proyecto.
    Ejemplo:
        Nota: Marynn Yahara Minetto Peralta es cliente de Prada y Beyond.
    """

    texto_nota = construir_texto_nota_multiproyecto(resumen_cliente)

    if texto_nota == "":
        return

    fila_nota = fila_total + 3

    # Descombinar una nota anterior si la hubiera en esa fila.
    for rango in list(ws.merged_cells.ranges):
        if rango.min_row == fila_nota and rango.max_row == fila_nota:
            ws.unmerge_cells(str(rango))

    # Limpiar zona de nota anterior.
    for col in range(1, 9):
        ws.cell(row=fila_nota, column=col).value = None

    # "Nota:" en negrita y cursiva.
    celda_titulo = ws.cell(row=fila_nota, column=1)
    celda_titulo.value = "Nota:"
    celda_titulo.font = Font(bold=True, italic=True)
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center")

    # Texto dinámico de la nota.
    celda_texto = ws.cell(row=fila_nota, column=2)
    celda_texto.value = texto_nota
    celda_texto.font = Font(bold=False, italic=False)
    celda_texto.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    # Combinar columnas para que la nota tenga espacio.
    ws.merge_cells(
        start_row=fila_nota,
        start_column=2,
        end_row=fila_nota,
        end_column=8,
    )

    ws.row_dimensions[fila_nota].height = 24

# Busca la plantilla Plantilla_Cobranza_por_cliente_para_gerencia.xlsx
def buscar_plantilla():
    ruta = CARPETA_PLANTILLAS / NOMBRE_PLANTILLA_GERENCIA

    if not ruta.exists():
        raise FileNotFoundError(
            "No se encontró la plantilla.\n"
            f"Ruta esperada: {ruta}"
        )

    return ruta

# Lee Base_Cobranzas_Ajustado.xlsx, convierte tipos y filtra proyectos vigentes.
def cargar_base_ajustada():
    ruta = SALIDA_LISTA / NOMBRE_BASE_AJUSTADA

    if not ruta.exists():
        raise FileNotFoundError(
            "No se encontró Base_Cobranzas_Ajustado.xlsx.\n"
            "Ejecuta primero Descarga_Cobranzas.py o usa EJECUTAR_DESCARGA = True.\n"
            f"Ruta esperada: {ruta}"
        )

    archivo = pd.ExcelFile(ruta, engine="openpyxl")

    if NOMBRE_HOJA_BASE_AJUSTADA not in archivo.sheet_names:
        raise ValueError(
            f"No se encontró la hoja {NOMBRE_HOJA_BASE_AJUSTADA} en {ruta.name}."
        )

    base = pd.read_excel(
        ruta,
        sheet_name=NOMBRE_HOJA_BASE_AJUSTADA,
        engine="openpyxl",
    )

    base = convertir_tipos_base(base)
    base = filtrar_proyectos_vigentes(base)

    print("✅ Base ajustada cargada:")
    print(ruta.resolve())

    return base, ruta

# Verifica que la base tenga las columnas necesarias para generar el reporte.
def validar_base(base):
    columnas_obligatorias = [
        "Proyecto",
        "Fecha_Programada",
        "Monto_Cuota_Pagado",
        "SaldoPorPagarCuota",
        "Estado",
        "TipoFinanciamiento",
    ]

    faltantes = [col for col in columnas_obligatorias if col not in base.columns]

    if faltantes:
        raise ValueError(
            "Faltan columnas obligatorias en Base_Ajustada:\n"
            + "\n".join(f"- {col}" for col in faltantes)
        )

    indices = obtener_indices_inmueble(base)

    if not indices:
        raise ValueError("No se encontraron columnas TipoInmueble_1, TipoInmueble_2, etc.")

    obtener_columna_cliente(base)


# =========================================================
# 4. CÁLCULOS DEL REPORTE
# =========================================================

# Construye la tabla principal por cliente y unidad inmobiliaria. 
# Calcula vendido, cobrado, por cobrar, vencido, deuda, tasa de morosidad y días de atraso.
def construir_resumen_cliente_unidad(base):
    validar_base(base)

    fecha_corte = obtener_fecha_corte()
    indices = obtener_indices_inmueble(base)
    columna_cliente = obtener_columna_cliente(base)

    grupos = {}

    for idx, row in base.iterrows():
        proyecto = obtener_nombre_proyecto(row.get("Proyecto", ""))
        cliente_id = row.get(columna_cliente, "")
        cliente_id = normalizar_texto(cliente_id)

        cliente_nombre = obtener_nombre_cliente(row)
        unidades = obtener_unidades_fila(row, indices)
        unidad_key = obtener_clave_unidad(unidades)
        unidad_label = obtener_label_unidad(unidades)
        financiamiento = obtener_financiamiento(row)

        clave_grupo = (
            normalizar_texto(proyecto),
            cliente_id,
            unidad_key,
        )

        if clave_grupo not in grupos:
            grupos[clave_grupo] = {
                "Proyecto": proyecto,
                "Cliente_ID": cliente_id,
                "Cliente": cliente_nombre,
                "Unidad Inmobiliaria": unidad_label,
                "Tipo de Financiamiento": financiamiento,
                "Monto Vendido": 0.0,
                "Monto Cobrado": 0.0,
                "Por Cobrar": 0.0,
                "Vencido": 0.0,
                "Deuda": 0.0,
                "Tasa de Morosidad": 0.0,
                "Días de Atraso": 0,
                "_precios_unicos": {},
                "_dias_atraso": [],
            }

        grupo = grupos[clave_grupo]

        if not grupo["Cliente"] or grupo["Cliente"] == "Sin cliente":
            grupo["Cliente"] = cliente_nombre

        if not grupo["Tipo de Financiamiento"] and financiamiento:
            grupo["Tipo de Financiamiento"] = financiamiento

        for unidad in unidades:
            if unidad["clave"] not in grupo["_precios_unicos"]:
                grupo["_precios_unicos"][unidad["clave"]] = unidad["precio"]

        monto_pagado = pd.to_numeric(row.get("Monto_Cuota_Pagado", 0), errors="coerce")
        monto_pagado = 0 if pd.isna(monto_pagado) else float(monto_pagado)

        saldo = pd.to_numeric(row.get("SaldoPorPagarCuota", 0), errors="coerce")
        saldo = 0 if pd.isna(saldo) else float(saldo)

        estado = normalizar_texto(row.get("Estado", ""))

        if monto_pagado > 0:
            grupo["Monto Cobrado"] += monto_pagado

        if saldo > 0 and "PENDIENTE" in estado:
            grupo["Por Cobrar"] += saldo

        if saldo > 0 and "VENCIDO" in estado:
            grupo["Vencido"] += saldo

            fecha_programada = pd.to_datetime(
                row.get("Fecha_Programada"),
                errors="coerce",
            )

            if pd.notna(fecha_programada):
                dias = max((fecha_corte - fecha_programada).days, 0)
                grupo["_dias_atraso"].append(dias)

    registros = []

    for _, grupo in grupos.items():
        monto_vendido = sum(grupo["_precios_unicos"].values())

        if monto_vendido == 0:
            monto_vendido = grupo["Monto Cobrado"] + grupo["Por Cobrar"] + grupo["Vencido"]

        grupo["Monto Vendido"] = monto_vendido

        # En este reporte gerencial, Deuda se entiende como deuda vencida.
        grupo["Deuda"] = grupo["Vencido"]

        if monto_vendido > 0:
            grupo["Tasa de Morosidad"] = grupo["Vencido"] / monto_vendido
        else:
            grupo["Tasa de Morosidad"] = 0

        if grupo["_dias_atraso"]:
            grupo["Días de Atraso"] = int(round(mean(grupo["_dias_atraso"]), 0))
        else:
            grupo["Días de Atraso"] = 0

        grupo.pop("_precios_unicos", None)
        grupo.pop("_dias_atraso", None)

        registros.append(grupo)

    resumen = pd.DataFrame(registros)

    if resumen.empty:
        columnas = [
            "Proyecto", "Cliente_ID", "Cliente", "Unidad Inmobiliaria",
            "Tipo de Financiamiento", "Monto Vendido", "Monto Cobrado",
            "Por Cobrar", "Vencido", "Deuda", "Tasa de Morosidad",
            "Días de Atraso",
        ]
        resumen = pd.DataFrame(columns=columnas)

    return resumen

# Calcula los días de atraso promedio por proyecto usando la base directa
def calcular_dias_atraso_por_proyecto_desde_base(base):
    """
    Calcula los días de atraso promedio por proyecto usando la misma lógica
    del Reporte_Cobranzas:
    """

    fecha_corte = obtener_fecha_corte()

    base_temp = base.copy()
    base_temp["Proyecto_Corto"] = base_temp["Proyecto"].apply(obtener_nombre_proyecto)

    estado = base_temp["Estado"].astype(str).apply(normalizar_texto)

    fecha_programada = pd.to_datetime(
        base_temp["Fecha_Programada"],
        errors="coerce",
    )

    saldo = pd.to_numeric(
        base_temp["SaldoPorPagarCuota"],
        errors="coerce",
    ).fillna(0)

    base_morosa = base_temp[
        estado.str.contains("VENCIDO", na=False)
        & fecha_programada.notna()
        & (fecha_programada <= fecha_corte)
        & (saldo > 0)
    ].copy()

    if base_morosa.empty:
        return {}

    base_morosa["Fecha_Programada"] = pd.to_datetime(
        base_morosa["Fecha_Programada"],
        errors="coerce",
    )

    base_morosa["Dias_Atraso_Calculo"] = (
        fecha_corte - base_morosa["Fecha_Programada"]
    ).dt.days.clip(lower=0)

    resumen_dias = (
        base_morosa
        .groupby("Proyecto_Corto", as_index=False)["Dias_Atraso_Calculo"]
        .mean()
    )

    resultado = {}

    for _, row in resumen_dias.iterrows():
        proyecto = str(row["Proyecto_Corto"]).strip()
        dias = row["Dias_Atraso_Calculo"]

        if pd.isna(dias):
            dias = 0

        resultado[proyecto] = int(round(dias, 0))

    return resultado

# Resume la información por proyecto: precio venta, cobrado, por cobrar, vencido, tasa de morosidad, 
# clientes morosos y días promedio
def construir_resumen_proyecto(resumen_cliente, base=None):
    if resumen_cliente.empty:
        return pd.DataFrame(columns=[
            "Proyecto",
            "Precio Venta",
            "Cobrado",
            "Por Cobrar",
            "Vencido",
            "Tasa Morosidad",
            "Cantidad de Clientes Morosos",
            "Días de Atraso Promedio",
        ])

    if base is not None:
        dias_por_proyecto = calcular_dias_atraso_por_proyecto_desde_base(base)
    else:
        dias_por_proyecto = {}

    proyectos = []

    for proyecto, temp in resumen_cliente.groupby("Proyecto"):
        precio_venta = float(temp["Monto Vendido"].sum())
        cobrado = float(temp["Monto Cobrado"].sum())
        por_cobrar = float(temp["Por Cobrar"].sum())
        vencido = float(temp["Vencido"].sum())

        temp_morosos = temp[temp["Vencido"] > 0].copy()
        clientes_morosos = temp_morosos["Cliente_ID"].nunique()

        # IMPORTANTE:
        # Si tenemos la base, usamos el promedio por cuota vencida,
        # igual que Reporte_Cobranzas.
        if base is not None:
            dias_promedio = dias_por_proyecto.get(str(proyecto).strip(), 0)
        else:
            if not temp_morosos.empty:
                dias_promedio = int(round(temp_morosos["Días de Atraso"].mean(), 0))
            else:
                dias_promedio = 0

        tasa_morosidad = vencido / precio_venta if precio_venta > 0 else 0

        proyectos.append({
            "Proyecto": proyecto,
            "Precio Venta": precio_venta,
            "Cobrado": cobrado,
            "Por Cobrar": por_cobrar,
            "Vencido": vencido,
            "Tasa Morosidad": tasa_morosidad,
            "Cantidad de Clientes Morosos": int(clientes_morosos),
            "Días de Atraso Promedio": int(dias_promedio),
        })

    resumen_proyecto = pd.DataFrame(proyectos)

    resumen_proyecto = resumen_proyecto.sort_values(
        by=["Vencido", "Por Cobrar", "Precio Venta"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    return resumen_proyecto


# =========================================================
# 5. UTILIDADES DE FORMATO EXCEL
# =========================================================

# Guarda el formato de una fila de la plantilla.
def capturar_estilo_fila(ws, fila, max_col):
    estilos = {
        "height": ws.row_dimensions[fila].height,
        "cells": {},
    }

    for col in range(1, max_col + 1):
        celda = ws.cell(row=fila, column=col)
        estilos["cells"][col] = {
            "style": copy(celda._style),
            "font": copy(celda.font),
            "fill": copy(celda.fill),
            "border": copy(celda.border),
            "alignment": copy(celda.alignment),
            "number_format": celda.number_format,
            "protection": copy(celda.protection),
        }

    return estilos

# Aplica un estilo guardado a una nueva fila.
def aplicar_estilo_fila(ws, fila, estilo, max_col):
    if estilo.get("height") is not None:
        ws.row_dimensions[fila].height = estilo["height"]

    for col in range(1, max_col + 1):
        celda = ws.cell(row=fila, column=col)
        cfg = estilo["cells"].get(col)
        if not cfg:
            continue

        celda._style = copy(cfg["style"])
        celda.font = copy(cfg["font"])
        celda.fill = copy(cfg["fill"])
        celda.border = copy(cfg["border"])
        celda.alignment = copy(cfg["alignment"])
        celda.number_format = cfg["number_format"]
        celda.protection = copy(cfg["protection"])

# Alterna estilos de filas: blanco, gris, blanco, gris.
def obtener_estilo_alternado(estilos, indice):
    """
    Devuelve el estilo de fila según el orden visual de la plantilla.
    """
    if not estilos:
        raise ValueError("No se enviaron estilos para alternar filas.")

    return estilos[indice % len(estilos)]

# Centra el contenido de una fila.
def centrar_rango_fila(ws, fila, col_inicio, col_fin):
    """Centra el contenido de una fila dinámica sin cambiar colores ni bordes."""
    for col in range(col_inicio, col_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

# Aplica alto fijo a las filas dinámicas de clientes.
def aplicar_alto_fila_cliente(ws, fila):
    """Aplica alto fijo a filas dinámicas de clientes: 21.6 pt ≈ 28.8 px."""
    ws.row_dimensions[fila].height = ALTO_FILA_CLIENTE_PUNTOS

# Limpia formatos condicionales heredados de la plantilla para evitar errores de Excel.
def limpiar_formato_condicional(ws):
    try:
        ws.conditional_formatting = ConditionalFormattingList()
    except Exception:
        pass

# Guarda el rango donde luego se aplicará una barra visual de deuda.
def aplicar_barra_deuda(ws, fila_inicio, fila_fin, max_deuda, columna='J'):
    if fila_fin < fila_inicio or max_deuda <= 0:
        return

    RANGOS_BARRAS_DEUDA.append({
        "hoja": ws.title,
        "rango": f"{columna}{fila_inicio}:{columna}{fila_fin}",
        "max_deuda": float(max_deuda),
    })

# Convierte color HEX a formato que entiende Excel COM
def color_excel_rgb(hex_color):
    """Convierte HEX RGB a entero BGR usado por Excel COM."""
    hex_color = str(hex_color).replace("#", "").strip()
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return r + (g * 256) + (b * 65536)

# Abre el archivo con Excel nativo y aplica las barras de deuda sin dañar el archivo.
def aplicar_barras_deuda_excel_nativo(ruta_excel):
    if not RANGOS_BARRAS_DEUDA:
        return

    try:
        import win32com.client as win32
    except ImportError as exc:
        raise ImportError(
            "No se pudo aplicar las barras sin aviso de Excel porque falta pywin32.\n"
            "Instálalo una sola vez con: pip install pywin32"
        ) from exc

    ruta_excel = Path(ruta_excel).resolve()

    excel = None
    wb_excel = None

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False

        wb_excel = excel.Workbooks.Open(str(ruta_excel))

        # Constantes Excel.
        xlConditionValueNumber = 0
        xlDataBarFillGradient = 1

        color_barra = color_excel_rgb("E73349")  # Rojo corporativo con degradado nativo de Excel

        for item in RANGOS_BARRAS_DEUDA:
            ws_excel = wb_excel.Worksheets(item["hoja"])
            rng = ws_excel.Range(item["rango"])

            # Borra solo las reglas previas del rango de deuda y crea una nueva
            # regla nativa válida para Excel.
            rng.FormatConditions.Delete()
            regla = rng.FormatConditions.AddDatabar()
            regla.ShowValue = True

            regla.MinPoint.Modify(xlConditionValueNumber, 0)
            regla.MaxPoint.Modify(xlConditionValueNumber, float(item["max_deuda"]))

            try:
                # 1 = degradado nativo de Excel. Evita que la barra quede de un solo color.
                regla.BarFillType = xlDataBarFillGradient
            except Exception:
                pass

            try:
                regla.BarColor.Color = color_barra
            except Exception:
                pass

        wb_excel.Save()

    finally:
        if wb_excel is not None:
            wb_excel.Close(SaveChanges=True)
        if excel is not None:
            excel.Quit()

# Descombina celdas desde cierta fila hacia abajo.
def descombinar_desde_fila(ws, fila_inicio):
    rangos = list(ws.merged_cells.ranges)

    for rango in rangos:
        if rango.min_row >= fila_inicio:
            ws.unmerge_cells(str(rango))

# Borra filas antiguas desde cierta fila.
def limpiar_desde_fila(ws, fila_inicio):
    descombinar_desde_fila(ws, fila_inicio)

    if ws.max_row >= fila_inicio:
        ws.delete_rows(fila_inicio, ws.max_row - fila_inicio + 1)

# Combina celdas solo si todavía no están combinadas.
def combinar_seguro(ws, rango):
    if rango not in [str(r) for r in ws.merged_cells.ranges]:
        ws.merge_cells(rango)

# Aplica combinaciones de celdas en la hoja Reporte_General
def aplicar_merges_reporte_general(ws, fila, tipo):
    if tipo == "proyecto":
        rangos = ["A:E", "F:G", "H:I", "J:K", "L:M"]
    elif tipo == "encabezado":
        rangos = ["B:E", "J:K"]
    else:
        rangos = ["B:E", "J:K"]

    for rango in rangos:
        inicio, fin = rango.split(":")
        combinar_seguro(ws, f"{inicio}{fila}:{fin}{fila}")

# Aplica combinaciones de celdas en la hoja Reporte_Cliente
def aplicar_merges_reporte_cliente(ws, fila):
    rangos = ["B:D", "E:F", "G:H", "I:J", "K:L", "O:P"]

    for rango in rangos:
        inicio, fin = rango.split(":")
        combinar_seguro(ws, f"{inicio}{fila}:{fin}{fila}")

# Aplica formato de monto, porcentaje y días en Reporte_General.
def aplicar_formatos_num_reporte_general(ws, fila):
    for col in [8, 9, 10]:
        ws.cell(row=fila, column=col).number_format = FORMATO_MONTO

    ws.cell(row=fila, column=12).number_format = FORMATO_PORCENTAJE
    ws.cell(row=fila, column=13).number_format = FORMATO_DIAS

# Aplica formato de monto, porcentaje y días en Reporte_Cliente.
def aplicar_formatos_num_reporte_cliente(ws, fila):
    for col in [9, 11]:
        ws.cell(row=fila, column=col).number_format = FORMATO_MONTO

    ws.cell(row=fila, column=13).number_format = FORMATO_DIAS
    ws.cell(row=fila, column=15).number_format = FORMATO_PORCENTAJE

# Aplica formato de monto, porcentaje y días en Reporte_Proyecto.
def aplicar_formatos_num_reporte_proyecto(ws, fila):
    for col in [2, 3, 4, 5]:
        ws.cell(row=fila, column=col).number_format = FORMATO_MONTO

    ws.cell(row=fila, column=6).number_format = FORMATO_PORCENTAJE
    ws.cell(row=fila, column=7).number_format = FORMATO_DIAS
    ws.cell(row=fila, column=8).number_format = FORMATO_DIAS

# Inserta la hoja Base_Ajustada dentro del reporte final y la oculta.
def escribir_base_ajustada_en_reporte(wb, base):
    if NOMBRE_HOJA_BASE_AJUSTADA in wb.sheetnames:
        ws_base = wb[NOMBRE_HOJA_BASE_AJUSTADA]
        wb.remove(ws_base)

    ws_base = wb.create_sheet(NOMBRE_HOJA_BASE_AJUSTADA)

    for fila in dataframe_to_rows(base, index=False, header=True):
        ws_base.append(fila)

    ws_base.freeze_panes = "A2"

    for col in range(1, ws_base.max_column + 1):
        ws_base.column_dimensions[get_column_letter(col)].width = 15

    encabezados = {celda.value: celda.column for celda in ws_base[1]}

    for columna in COLUMNAS_FECHA:
        if columna in encabezados:
            col_idx = encabezados[columna]
            for fila in range(2, ws_base.max_row + 1):
                ws_base.cell(row=fila, column=col_idx).number_format = "dd/mm/yyyy"

    for columna, col_idx in encabezados.items():
        if es_columna_numero(columna):
            for fila in range(2, ws_base.max_row + 1):
                ws_base.cell(row=fila, column=col_idx).number_format = "#,##0.00"

    ws_base.sheet_state = "hidden"


# =========================================================
# 6. LLENADO DE HOJAS
# =========================================================

# Llena la hoja Reporte_General, agrupando clientes morosos por proyecto.
def llenar_reporte_general(wb, resumen_cliente):
    ws = wb[HOJA_REPORTE_GENERAL]

    estilo_proyecto = capturar_estilo_fila(ws, 7, 13)
    estilo_encabezado = capturar_estilo_fila(ws, 8, 13)

    # La plantilla tiene filas de datos alternadas:
    estilo_dato_blanco = capturar_estilo_fila(ws, 9, 13)
    estilo_dato_gris = capturar_estilo_fila(ws, 10, 13) if ws.max_row >= 10 else estilo_dato_blanco
    estilos_datos = [estilo_dato_blanco, estilo_dato_gris]

    estilo_blanco = capturar_estilo_fila(ws, 15, 13) if ws.max_row >= 15 else estilo_dato_blanco

    limpiar_desde_fila(ws, 7)
    limpiar_formato_condicional(ws)

    deudores = resumen_cliente[resumen_cliente["Deuda"] > 0].copy()
    deudores = deudores.sort_values(
        by=["Proyecto", "Deuda", "Días de Atraso"],
        ascending=[True, False, False],
    )

    # El resumen superior debe cuadrar con los bloques por proyecto..
    total_deuda_general = 0.0
    total_clientes_general = 0

    ws["C5"] = 0
    ws["G5"] = 0
    ws["C5"].number_format = FORMATO_MONTO
    ws["G5"].number_format = FORMATO_DIAS

    proyectos_disponibles = resumen_cliente["Proyecto"].dropna().unique().tolist()

    proyectos_ordenados = [
        p for p in ORDEN_PROYECTOS
        if p in proyectos_disponibles
    ] + [
        p for p in proyectos_disponibles
        if p not in ORDEN_PROYECTOS
    ]

    fila_actual = 7

    for proyecto in proyectos_ordenados:
        temp = deudores[deudores["Proyecto"] == proyecto].copy()
        temp = temp.sort_values(
            by=["Deuda", "Días de Atraso", "Monto Vendido"],
            ascending=[False, False, False],
        ).reset_index(drop=True)

        deuda_proyecto = float(temp["Deuda"].sum()) if not temp.empty else 0.0
        clientes_proyecto = int(temp["Cliente_ID"].nunique()) if not temp.empty else 0

        # Estos son los mismos valores que se muestran en el encabezado de cada proyecto.
        # El resumen superior se calcula sumando exactamente estos importes/contadores.
        total_deuda_general += deuda_proyecto
        total_clientes_general += clientes_proyecto

        aplicar_estilo_fila(ws, fila_actual, estilo_proyecto, 13)
        aplicar_merges_reporte_general(ws, fila_actual, "proyecto")

        ws.cell(row=fila_actual, column=1).value = normalizar_texto(proyecto)
        ws.cell(row=fila_actual, column=6).value = "DEUDA TOTAL"
        ws.cell(row=fila_actual, column=8).value = deuda_proyecto
        ws.cell(row=fila_actual, column=10).value = "CLIENTE"
        ws.cell(row=fila_actual, column=12).value = clientes_proyecto
        ws.cell(row=fila_actual, column=8).number_format = FORMATO_MONTO
        centrar_rango_fila(ws, fila_actual, 1, 13)

        fila_encabezado = fila_actual + 1
        aplicar_estilo_fila(ws, fila_encabezado, estilo_encabezado, 13)
        aplicar_merges_reporte_general(ws, fila_encabezado, "encabezado")

        encabezados = {
            1: "Rank",
            2: "Cliente",
            6: "Unidad",
            7: "Financiamiento",
            8: "Precio Venta",
            9: "Monto Cobrado",
            10: "Monto Deuda",
            12: "Tasa de Morosidad",
            13: "Días de atraso",
        }

        for col, valor in encabezados.items():
            ws.cell(row=fila_encabezado, column=col).value = valor

        centrar_rango_fila(ws, fila_encabezado, 1, 13)

        if temp.empty:
            filas_datos = 1
        else:
            filas_datos = len(temp)

        for i in range(filas_datos):
            fila = fila_encabezado + 1 + i
            aplicar_estilo_fila(ws, fila, obtener_estilo_alternado(estilos_datos, i), 13)
            aplicar_merges_reporte_general(ws, fila, "dato")

            if temp.empty:
                ws.cell(row=fila, column=1).value = None
                ws.cell(row=fila, column=2).value = ""
                ws.cell(row=fila, column=6).value = ""
                ws.cell(row=fila, column=7).value = ""
                ws.cell(row=fila, column=8).value = 0
                ws.cell(row=fila, column=9).value = 0
                ws.cell(row=fila, column=10).value = 0
                ws.cell(row=fila, column=12).value = 0
                ws.cell(row=fila, column=13).value = 0
            else:
                row = temp.iloc[i]
                ws.cell(row=fila, column=1).value = i + 1
                ws.cell(row=fila, column=2).value = row["Cliente"]
                ws.cell(row=fila, column=6).value = row["Unidad Inmobiliaria"]
                ws.cell(row=fila, column=7).value = row["Tipo de Financiamiento"]
                ws.cell(row=fila, column=8).value = float(row["Monto Vendido"])
                ws.cell(row=fila, column=9).value = float(row["Monto Cobrado"])
                ws.cell(row=fila, column=10).value = float(row["Deuda"])
                ws.cell(row=fila, column=12).value = float(row["Tasa de Morosidad"])
                ws.cell(row=fila, column=13).value = int(row["Días de Atraso"])

            aplicar_formatos_num_reporte_general(ws, fila)
            centrar_rango_fila(ws, fila, 1, 13)
            aplicar_alto_fila_cliente(ws, fila)

        if not temp.empty:
            fila_inicio_barras = fila_encabezado + 1
            fila_fin_barras = fila_encabezado + filas_datos
            max_deuda_bloque = float(temp["Deuda"].max())
            aplicar_barra_deuda(ws, fila_inicio_barras, fila_fin_barras, max_deuda_bloque, columna="J")

        fila_actual = fila_encabezado + 1 + filas_datos

        for _ in range(2):
            aplicar_estilo_fila(ws, fila_actual, estilo_blanco, 13)
            fila_actual += 1

    # Resumen superior: suma exacta de lo mostrado en cada bloque de proyecto.
    ws["C5"] = total_deuda_general
    ws["G5"] = total_clientes_general
    ws["C5"].number_format = FORMATO_MONTO
    ws["G5"].number_format = FORMATO_DIAS

    ws.freeze_panes = "A8"

# Llena la hoja Reporte_Cliente, mostrando el ranking general de clientes con deuda.
def llenar_reporte_cliente(wb, resumen_cliente):
    ws = wb[HOJA_REPORTE_CLIENTE]

    # La plantilla tiene filas alternadas: fila 5 blanco, fila 6 gris #F4F6F8.
    estilo_dato_blanco = capturar_estilo_fila(ws, 5, 16)
    estilo_dato_gris = capturar_estilo_fila(ws, 6, 16) if ws.max_row >= 6 else estilo_dato_blanco
    estilos_datos = [estilo_dato_blanco, estilo_dato_gris]

    limpiar_desde_fila(ws, 5)
    limpiar_formato_condicional(ws)

    deudores = resumen_cliente[resumen_cliente["Deuda"] > 0].copy()
    deudores = deudores.sort_values(
        by=["Deuda", "Días de Atraso", "Monto Vendido"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    filas_datos = len(deudores) if not deudores.empty else 1

    for i in range(filas_datos):
        fila = 5 + i
        aplicar_estilo_fila(ws, fila, obtener_estilo_alternado(estilos_datos, i), 16)
        aplicar_merges_reporte_cliente(ws, fila)

        if deudores.empty:
            ws.cell(row=fila, column=1).value = None
            ws.cell(row=fila, column=2).value = ""
            ws.cell(row=fila, column=5).value = ""
            ws.cell(row=fila, column=7).value = ""
            ws.cell(row=fila, column=9).value = 0
            ws.cell(row=fila, column=11).value = 0
            ws.cell(row=fila, column=13).value = 0
            ws.cell(row=fila, column=14).value = ""
            ws.cell(row=fila, column=15).value = 0
        else:
            row = deudores.iloc[i]
            ws.cell(row=fila, column=1).value = i + 1
            ws.cell(row=fila, column=2).value = row["Cliente"]
            ws.cell(row=fila, column=5).value = row["Proyecto"]
            ws.cell(row=fila, column=7).value = row["Tipo de Financiamiento"]
            ws.cell(row=fila, column=9).value = float(row["Monto Vendido"])
            ws.cell(row=fila, column=11).value = float(row["Deuda"])
            ws.cell(row=fila, column=13).value = int(row["Días de Atraso"])
            ws.cell(row=fila, column=14).value = row["Unidad Inmobiliaria"]
            ws.cell(row=fila, column=15).value = float(row["Tasa de Morosidad"])

        aplicar_formatos_num_reporte_cliente(ws, fila)
        centrar_rango_fila(ws, fila, 1, 16)
        aplicar_alto_fila_cliente(ws, fila)

    if not deudores.empty:
        fila_inicio_barras = 5
        fila_fin_barras = 5 + filas_datos - 1
        max_deuda_cliente = float(deudores["Deuda"].max())
        aplicar_barra_deuda(ws, fila_inicio_barras, fila_fin_barras, max_deuda_cliente, columna="K")

    ws.freeze_panes = "A5"


# Crea una fórmula para validar que Precio Venta = Cobrado + Por Cobrar + Vencido.
def formula_validador_reporte_proyecto(fila):
    """
    Valida que el precio de venta cuadre con:
    Cobrado + Por Cobrar + Vencido.
    """
    return f'=IF(ROUND(B{fila},0)=ROUND(C{fila}+D{fila}+E{fila},0),"OK","REVISAR")'

# Llena la hoja Reporte_Proyecto con el resumen por proyecto y agrega la nota multiproyecto debajo del TOTAL.
def llenar_reporte_proyecto(wb, resumen_proyecto, resumen_cliente):
    ws = wb[HOJA_REPORTE_PROYECTO]

    # La plantilla tiene filas alternadas: fila 5 blanco, fila 6 gris #F4F6F8.
    estilo_dato_blanco = capturar_estilo_fila(ws, 5, 9)
    estilo_dato_gris = capturar_estilo_fila(ws, 6, 9) if ws.max_row >= 6 else estilo_dato_blanco
    estilos_datos = [estilo_dato_blanco, estilo_dato_gris]
    estilo_total = capturar_estilo_fila(ws, 8, 9)

    limpiar_desde_fila(ws, 5)

    resumen_proyecto = resumen_proyecto.copy()
    resumen_proyecto = resumen_proyecto.sort_values(
        by=["Vencido", "Por Cobrar", "Precio Venta"],
        ascending=[False, False, False],
    ).reset_index(drop=True)

    fila_inicio = 5

    for i, row in resumen_proyecto.iterrows():
        fila = fila_inicio + i
        aplicar_estilo_fila(ws, fila, obtener_estilo_alternado(estilos_datos, i), 9)

        ws.cell(row=fila, column=1).value = row["Proyecto"]
        ws.cell(row=fila, column=2).value = float(row["Precio Venta"])
        ws.cell(row=fila, column=3).value = float(row["Cobrado"])
        ws.cell(row=fila, column=4).value = float(row["Por Cobrar"])
        ws.cell(row=fila, column=5).value = float(row["Vencido"])
        ws.cell(row=fila, column=6).value = float(row["Tasa Morosidad"])
        ws.cell(row=fila, column=7).value = int(row["Cantidad de Clientes Morosos"])
        ws.cell(row=fila, column=8).value = int(row["Días de Atraso Promedio"])
        ws.cell(row=fila, column=9).value = formula_validador_reporte_proyecto(fila)

        aplicar_formatos_num_reporte_proyecto(ws, fila)
        centrar_rango_fila(ws, fila, 1, 9)

    fila_total = fila_inicio + len(resumen_proyecto)
    aplicar_estilo_fila(ws, fila_total, estilo_total, 9)

    total_precio = float(resumen_proyecto["Precio Venta"].sum()) if not resumen_proyecto.empty else 0.0
    total_cobrado = float(resumen_proyecto["Cobrado"].sum()) if not resumen_proyecto.empty else 0.0
    total_por_cobrar = float(resumen_proyecto["Por Cobrar"].sum()) if not resumen_proyecto.empty else 0.0
    total_vencido = float(resumen_proyecto["Vencido"].sum()) if not resumen_proyecto.empty else 0.0

    # Total dinámico de clientes morosos:
    # se suma lo mostrado por cada proyecto.
    # Esto evita confusión cuando un mismo cliente aparece en más de un proyecto.
    total_clientes_morosos = int(
        resumen_proyecto["Cantidad de Clientes Morosos"].fillna(0).sum()
    ) if not resumen_proyecto.empty else 0

    # Total de días de atraso:
    # mantiene el criterio global del resumen cliente.
    # Los días por proyecto ya vienen calculados desde la base para coincidir con Reporte_Cobranzas.
    deudores = resumen_cliente[resumen_cliente["Deuda"] > 0].copy()

    if not deudores.empty:
        dias_total = int(round(deudores["Días de Atraso"].mean(), 0))
    else:
        dias_total = 0

    tasa_total = total_vencido / total_precio if total_precio > 0 else 0

    ws.cell(row=fila_total, column=1).value = "TOTAL"
    ws.cell(row=fila_total, column=2).value = total_precio
    ws.cell(row=fila_total, column=3).value = total_cobrado
    ws.cell(row=fila_total, column=4).value = total_por_cobrar
    ws.cell(row=fila_total, column=5).value = total_vencido
    ws.cell(row=fila_total, column=6).value = tasa_total
    ws.cell(row=fila_total, column=7).value = total_clientes_morosos
    ws.cell(row=fila_total, column=8).value = dias_total
    ws.cell(row=fila_total, column=9).value = formula_validador_reporte_proyecto(fila_total)

    aplicar_formatos_num_reporte_proyecto(ws, fila_total)
    centrar_rango_fila(ws, fila_total, 1, 9)

    # Nota dinámica debajo de la fila TOTAL.
    agregar_nota_multiproyecto_reporte_proyecto(
        ws=ws,
        resumen_cliente=resumen_cliente,
        fila_total=fila_total,
    )

    ws.freeze_panes = "A5"
# # generar_reporte_gerencia
def generar_reporte_gerencia(base):
    RANGOS_BARRAS_DEUDA.clear()

    ruta_plantilla = buscar_plantilla()
    ruta_salida = SALIDA_LISTA / NOMBRE_REPORTE_GERENCIA

    resumen_cliente = construir_resumen_cliente_unidad(base)

    # Se envía también la base para que los días de atraso por proyecto
    # se calculen igual que en Reporte_Cobranzas.
    resumen_proyecto = construir_resumen_proyecto(resumen_cliente, base)

    wb = load_workbook(ruta_plantilla)

    hojas_necesarias = [
        HOJA_REPORTE_GENERAL,
        HOJA_REPORTE_CLIENTE,
        HOJA_REPORTE_PROYECTO,
    ]

    faltantes = [hoja for hoja in hojas_necesarias if hoja not in wb.sheetnames]

    if faltantes:
        raise ValueError(
            "La plantilla no tiene las hojas necesarias:\n"
            + "\n".join(f"- {hoja}" for hoja in faltantes)
        )

    escribir_base_ajustada_en_reporte(wb, base)
    llenar_reporte_general(wb, resumen_cliente)
    llenar_reporte_cliente(wb, resumen_cliente)
    llenar_reporte_proyecto(wb, resumen_proyecto, resumen_cliente)

    # Limpieza final de vínculos externos, sin tocar las barras de datos.
    try:
        wb._external_links = []
    except Exception:
        pass

    try:
        if ruta_salida.exists():
            ruta_salida.unlink()
    except PermissionError:
        raise PermissionError(
            f"No se pudo reemplazar {ruta_salida.name}. "
            "Cierra el archivo si está abierto en Excel y vuelve a ejecutar."
        )

    wb.save(ruta_salida)

    # Aplicar barras de datos con Excel nativo después de guardar.
    # Esto mantiene las barritas sin generar el aviso de reparación de contenido.
    aplicar_barras_deuda_excel_nativo(ruta_salida)

    print("====================================")
    print("REPORTE PARA GERENCIA GENERADO")
    print("====================================")
    print("✅ Archivo generado:")
    print(ruta_salida.resolve())
    return ruta_salida

# Ejecuta la descarga si corresponde, carga la base ajustada y genera el reporte.
def ejecutar_reporte_gerencia(actualizar_base=True):
    if actualizar_base:
        ejecutar_descarga_cobranzas()

    base, ruta_base = cargar_base_ajustada()
    ruta_reporte = generar_reporte_gerencia(base)

    return base, ruta_base, ruta_reporte


def main():
    ejecutar_reporte_gerencia(actualizar_base=EJECUTAR_DESCARGA)


if __name__ == "__main__":
    main()