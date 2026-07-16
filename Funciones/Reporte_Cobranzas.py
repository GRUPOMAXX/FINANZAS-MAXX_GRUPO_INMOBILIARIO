# Programa para descargar la base de datos de cobranzas en Evolta

# Creado por Eduardo Miguel Huamani Acosta                            15/07/26


# Este automatización trata de elaborar un reporte de cobranzas por proyecto. Al momento de elaborar el reporte se tendría que
# formular cada proyecto en una hoja, revisar que la información esté hasta el cierre de mes, demandando horas de trabajo.

# Este código permite extraer información de las cobranzas y elabora un reporte en base al cierre de mes, si en caso se
# encuentra alguna cobranza o venta, solo se tomará hasta el cierre de mes. Este reporte de lo que se hacía en 2 - 3 horas, 
# con el codigo se podrá tener el reporte completo en 2 - 3 minutos.

# =============================================================================================================================

# LIBRERIAS 

import sys
import re
import unicodedata
from pathlib import Path
from copy import copy

import pandas as pd                                  
from openpyxl import load_workbook                  
from openpyxl.utils import get_column_letter        
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill


BASE_DIR = Path(__file__).resolve().parent.parent

if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from Conexiones.connection import ( FECHA_CORTE_REPORTE, ANIO_MENSUAL_REPORTE, PROYECTO,)

try:
    from Conexiones.connection import PROYECTOS_VIGENTES_COBRANZAS  # noqa: E402
except ImportError:
    PROYECTOS_VIGENTES_COBRANZAS = []

import Descarga_Cobranzas as descarga  # noqa: E402


# =========================================================
# 1. VARIABLES DEL REPORTE
# =========================================================

EJECUTAR_DESCARGA = True

SALIDA_LISTA = BASE_DIR / "Flujo" / "Output"

NOMBRE_HOJA_PLANTILLA = "Prada"
NOMBRE_HOJA_BASE = "Base_Cobranzas"
NOMBRE_REPORTE_COBRANZAS = "Reporte_Cobranzas.xlsx"
NOMBRE_BASE_AJUSTADA = "Base_Cobranzas_Ajustado.xlsx"
NOMBRE_HOJA_BASE_AJUSTADA = "Base_Ajustada"

RUTAS_PLANTILLA_REPORTE = [
    BASE_DIR / "Flujo" / "Input" / "Plantillas" / "Plantilla de reporte.xlsx"
]

FORMATO_MONTO_REPORTE = '#,##0;-#,##0;-'
FORMATO_PORCENTAJE_REPORTE = '0.0%'
FORMATO_DIAS_REPORTE = '#,##0;-#,##0;-'

COLOR_TOTAL_DETALLE = 'F2F2F2'
COLOR_TOTAL_FILA = 'D0D0D0'


SALIDA_LISTA.mkdir(parents=True, exist_ok=True)
RUTAS_PLANTILLA_REPORTE[0].parent.mkdir(parents=True, exist_ok=True)

COLUMNAS_FECHA = [
    "Fecha_Programada",
    "FechaPago",
]

COLUMNAS_NUMERO_FIJAS = [
    "Monto_Cuota",
    "Monto_Cuota_Pagado",
    "SaldoPorPagarCuota",
]

COLUMNAS_NUMERO_CON_INDICE = [
    "PrecioLista",
    "PrecioVenta",
]

COLUMNAS_OBLIGATORIAS_REPORTE = [
    "Proyecto",
    "Fecha_Programada",
    "FechaPago",
    "Monto_Cuota",
    "Monto_Cuota_Pagado",
    "SaldoPorPagarCuota",
    "TipoFinanciamiento",
    "Estado",
]


# =========================================================
# 2. FUNCIONES GENERALES
# =========================================================

# Ejecuta el script Descargas_Cobranzas
def ejecutar_descarga_cobranzas():
    descarga.main()

# Limpia el texto para que se pueda comparar
def normalizar_texto(texto):
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(
        caracter for caracter in texto
        if not unicodedata.combining(caracter)
    )
    return texto

# Cambia el nombre largo por el estandar
def obtener_nombre_proyecto(proyecto):
    proyecto_normalizado = normalizar_texto(proyecto)

    for nombre_original, nombre_corto in PROYECTO.items():
        if normalizar_texto(nombre_original) == proyecto_normalizado:
            return nombre_corto

    return str(proyecto).strip()

# Limpia la columna que esta en letra y lo convierte en numero
def convertir_columna_numero(serie):
    serie = serie.astype(str).str.strip()
    serie = serie.str.replace("S/", "", regex=False)
    serie = serie.str.replace("US$", "", regex=False)
    serie = serie.str.replace("$", "", regex=False)
    serie = serie.str.replace(" ", "", regex=False)
    serie = serie.str.replace(",", "", regex=False)
    serie = serie.replace("", pd.NA)
    return pd.to_numeric(serie, errors="coerce")

#  Identifica que columna se debe de tratarse como numero
def es_columna_numero(columna):
    if columna in COLUMNAS_NUMERO_FIJAS:
        return True

    for campo in COLUMNAS_NUMERO_CON_INDICE:
        if re.match(rf"^{campo}_(\d+)$", str(columna)):
            return True

    return False

# Convierte fechas y montos a los formatos correctos
def convertir_tipos_base(base):
    base = base.copy()

    for columna in COLUMNAS_FECHA:
        if columna in base.columns:
            base[columna] = pd.to_datetime(
                base[columna],
                errors="coerce",
                dayfirst=True,
            )

    for columna in base.columns:
        if es_columna_numero(columna):
            base[columna] = convertir_columna_numero(base[columna])

    return base

# Ajusta el nombre del proyecto para que entre como nombre en las hojas de Excel
def nombre_hoja_valido(nombre):
    nombre = str(nombre).strip()
    nombre = re.sub(r"[\[\]\*\?/\\:]", "-", nombre)
    nombre = nombre[:31]
    return nombre if nombre else "Proyecto"


# =========================================================
# 3. FECHA DE CORTE
# =========================================================

# Lee la fecha de corte del script Connection
def obtener_fecha_corte_reporte():
    if FECHA_CORTE_REPORTE is None or str(FECHA_CORTE_REPORTE).strip() == "":
        return None

    fecha_corte = pd.to_datetime(
        FECHA_CORTE_REPORTE,
        errors="coerce",
        dayfirst=True,
    )

    if pd.isna(fecha_corte):
        raise ValueError(
            "FECHA_CORTE_REPORTE no tiene un formato válido. "
            "Usa un formato como '30/06/2026'."
        )

    return fecha_corte

# Calcula el primer día del mes siguiente al corte
def obtener_inicio_mes_siguiente_corte():
    fecha_corte = obtener_fecha_corte_reporte()

    if fecha_corte is None:
        return None

    inicio_mes_corte = pd.Timestamp(
        year=fecha_corte.year,
        month=fecha_corte.month,
        day=1,
    )

    return inicio_mes_corte + pd.DateOffset(months=1)

# Devuelve el mes siguiente y el mes posterior, esto se utiliza para manejar rangos mensuales
def obtener_rango_mes_siguiente_corte():
    inicio_mes_siguiente = obtener_inicio_mes_siguiente_corte()

    if inicio_mes_siguiente is None:
        return None, None

    inicio_mes_posterior = inicio_mes_siguiente + pd.DateOffset(months=1)
    return inicio_mes_siguiente, inicio_mes_posterior


# =========================================================
# 4. CARGA DE BASE Y PLANTILLA
# =========================================================

# Busca la plantilla que se habia colocado en la carpeta Plantillas
def buscar_plantilla_reporte():
    for ruta in RUTAS_PLANTILLA_REPORTE:
        if ruta.exists():
            return ruta

    ruta = RUTAS_PLANTILLA_REPORTE[0]
    raise FileNotFoundError(
        "❌ No se encontró la plantilla del reporte de cobranzas.\n"
        "Coloca el archivo 'Plantilla de reporte.xlsx' en esta ruta:\n"
        f"- {ruta}"
    )

# Carga la base de datos que se había adaptado con corte al cierre de mes
def cargar_base_consolidada():
    """
    Carga únicamente la base ajustada que usará el reporte.

    Importante:
    El ajuste de fecha de cierre, estados y saldos negativos ya debe venir
    resuelto desde Descarga_Cobranzas.py. Este reporte no vuelve a modificar
    Estado, FechaPago, Monto_Cuota_Pagado ni SaldoPorPagarCuota.
    """
    ruta_ajustada = SALIDA_LISTA / NOMBRE_BASE_AJUSTADA

    if not ruta_ajustada.exists():
        raise FileNotFoundError(
            "❌ No se encontró Base_Cobranzas_Ajustado.xlsx.\n"
            "Primero ejecuta Descarga_Cobranzas.py para generar la base ajustada.\n"
            f"Ruta esperada: {ruta_ajustada}"
        )

    wb_info = pd.ExcelFile(ruta_ajustada, engine="openpyxl")

    if NOMBRE_HOJA_BASE_AJUSTADA not in wb_info.sheet_names:
        raise ValueError(
            "❌ El archivo Base_Cobranzas_Ajustado.xlsx no tiene la hoja Base_Ajustada.\n"
            f"Hojas encontradas: {', '.join(wb_info.sheet_names)}"
        )

    base = pd.read_excel(
        ruta_ajustada,
        sheet_name=NOMBRE_HOJA_BASE_AJUSTADA,
        engine="openpyxl",
    )

    base = convertir_tipos_base(base)

    return base, ruta_ajustada

# Verifica que la base de datos tenga las columna minimas para generar el reporte
def validar_columnas_reporte(base):
    faltantes = [col for col in COLUMNAS_OBLIGATORIAS_REPORTE if col not in base.columns]

    if faltantes:
        raise ValueError(
            "❌ Faltan columnas obligatorias para generar el Reporte_Cobranzas:\n"
            + "\n".join(f"- {col}" for col in faltantes)
        )

    columnas_tipo = obtener_columnas_tipo_inmueble({col: "" for col in base.columns})
    columnas_nro = [col for col in base.columns if re.match(r"^NroInmueble_(\d+)$", str(col))]

    if not columnas_tipo:
        raise ValueError("❌ No se encontraron columnas TipoInmueble_1, TipoInmueble_2, etc.")

    if not columnas_nro:
        print("⚠️ No se encontraron columnas NroInmueble_N. Las unidades se calcularán con limitaciones.")

# Identifica que proyectos debe de tener como nombre en la hoja de Excel
def obtener_proyectos_reporte(base):
    if "Proyecto" not in base.columns:
        raise ValueError("No existe la columna Proyecto en la base consolidada.")

    proyectos_base = (
        base["Proyecto"]
        .dropna()
        .astype(str)
        .map(obtener_nombre_proyecto)
        .str.strip()
    )

    proyectos_base = [p for p in proyectos_base.unique().tolist() if p]

    orden_predefinido = list(PROYECTO.values())
    proyectos_ordenados = [p for p in orden_predefinido if p in proyectos_base]
    otros = [p for p in proyectos_base if p not in proyectos_ordenados]

    proyectos_finales = proyectos_ordenados + otros

    if PROYECTOS_VIGENTES_COBRANZAS:
        proyectos_solicitados = [
            normalizar_texto(proyecto)
            for proyecto in PROYECTOS_VIGENTES_COBRANZAS
        ]

        proyectos_finales = [
            proyecto
            for proyecto in proyectos_finales
            if normalizar_texto(proyecto) in proyectos_solicitados
        ]

    if not proyectos_finales:
        raise ValueError(
            "No se encontraron proyectos válidos para generar el Reporte_Cobranzas. "
            "Revisa PROYECTOS_VIGENTES_COBRANZAS en Conexiones/connection.py."
        )

    return proyectos_finales


# =========================================================
# 5. PERIODOS DINÁMICOS
# =========================================================

# Convierte cualquier fecha al primer dia de su mes
def primer_dia_mes(fecha):
    fecha = pd.to_datetime(fecha, errors="coerce")

    if pd.isna(fecha):
        return None

    return pd.Timestamp(year=fecha.year, month=fecha.month, day=1)

# Genera una lista mensual entre dos fechas
def lista_meses(fecha_inicio, fecha_fin):
    fecha_inicio = primer_dia_mes(fecha_inicio)
    fecha_fin = primer_dia_mes(fecha_fin)

    if fecha_inicio is None or fecha_fin is None or fecha_inicio > fecha_fin:
        return []

    meses = []
    fecha = fecha_inicio

    while fecha <= fecha_fin:
        meses.append(fecha)
        fecha = fecha + pd.DateOffset(months=1)

    return meses

# Construye los años acumulados de la cobranza pasada
def construir_periodos_por_anio_reporte(fecha_inicio, fecha_fin):
    fecha_inicio = pd.to_datetime(fecha_inicio, errors="coerce")
    fecha_fin = pd.to_datetime(fecha_fin, errors="coerce")

    if pd.isna(fecha_inicio) or pd.isna(fecha_fin) or fecha_inicio > fecha_fin:
        return []

    periodos = []

    ultimo_anio_anual = min(fecha_fin.year, ANIO_MENSUAL_REPORTE - 1)

    for anio in range(fecha_inicio.year, ultimo_anio_anual + 1):
        periodos.append({"tipo": "anio", "valor": anio})

    if fecha_fin.year >= ANIO_MENSUAL_REPORTE:
        fecha_inicio_mensual = max(
            primer_dia_mes(fecha_inicio),
            pd.Timestamp(year=ANIO_MENSUAL_REPORTE, month=1, day=1),
        )

        for mes in lista_meses(fecha_inicio_mensual, fecha_fin):
            periodos.append({"tipo": "mes", "valor": mes})

    return periodos

# Construye periodos mensuales en el año del cierre de mes 
def construir_periodos_solo_mensual(fecha_inicio, fecha_fin):
    fecha_inicio = pd.to_datetime(fecha_inicio, errors="coerce")
    fecha_fin = pd.to_datetime(fecha_fin, errors="coerce")

    if pd.isna(fecha_inicio) or pd.isna(fecha_fin) or fecha_inicio > fecha_fin:
        return []

    return [
        {"tipo": "mes", "valor": mes}
        for mes in lista_meses(fecha_inicio, fecha_fin)
    ]

#  Calcula los periodos que tendrá cada bloque del proyecto: real, proyectado y moroso
def obtener_periodos_proyecto(base, proyecto):
    base_temp = base.copy()
    base_temp["Proyecto_Corto"] = base_temp["Proyecto"].apply(obtener_nombre_proyecto)
    base_proyecto = base_temp[base_temp["Proyecto_Corto"] == proyecto].copy()

    estado = base_proyecto["Estado"].astype(str).apply(normalizar_texto)
    estado_pendiente = estado.str.contains("PENDIENTE", na=False)
    estado_vencido = estado.str.contains("VENCIDO", na=False)

    fecha_corte = obtener_fecha_corte_reporte()
    inicio_mes_siguiente = obtener_inicio_mes_siguiente_corte()

    base_real = base_proyecto[
        (base_proyecto["Monto_Cuota_Pagado"].fillna(0) > 0)
        & (base_proyecto["FechaPago"].notna())
    ]

    if inicio_mes_siguiente is not None:
        base_proyectada = base_proyecto[
            estado_pendiente
            & (base_proyecto["Fecha_Programada"].notna())
            & (base_proyecto["Fecha_Programada"] >= inicio_mes_siguiente)
            & (base_proyecto["SaldoPorPagarCuota"].fillna(0) > 0)
        ]
    else:
        base_proyectada = base_proyecto[
            estado_pendiente
            & (base_proyecto["Fecha_Programada"].notna())
            & (base_proyecto["SaldoPorPagarCuota"].fillna(0) > 0)
        ]

    if fecha_corte is not None:
        base_morosa = base_proyecto[
            estado_vencido
            & (base_proyecto["Fecha_Programada"].notna())
            & (base_proyecto["Fecha_Programada"] <= fecha_corte)
            & (base_proyecto["SaldoPorPagarCuota"].fillna(0) > 0)
        ]
    else:
        base_morosa = base_proyecto[
            estado_vencido
            & (base_proyecto["Fecha_Programada"].notna())
            & (base_proyecto["SaldoPorPagarCuota"].fillna(0) > 0)
        ]

    if not base_real.empty:
        periodos_real = construir_periodos_por_anio_reporte(
            base_real["FechaPago"].min(),
            base_real["FechaPago"].max(),
        )
    else:
        periodos_real = []

    if not base_proyectada.empty:
        fecha_inicio_proy = (
            inicio_mes_siguiente
            if inicio_mes_siguiente is not None
            else base_proyectada["Fecha_Programada"].min()
        )

        periodos_proy = construir_periodos_solo_mensual(
            fecha_inicio_proy,
            base_proyectada["Fecha_Programada"].max(),
        )
    else:
        periodos_proy = []

    if not base_morosa.empty:
        fecha_fin_morosa = (
            fecha_corte
            if fecha_corte is not None
            else base_morosa["Fecha_Programada"].max()
        )

        periodos_morosa = construir_periodos_solo_mensual(
            base_morosa["Fecha_Programada"].min(),
            fecha_fin_morosa,
        )
    else:
        periodos_morosa = []

    return periodos_real, periodos_proy, periodos_morosa


# =========================================================
# 6. ESTILOS EXCEL
# =========================================================

# Copia formato de una celda a otra
def copiar_estilo(origen, destino):
    if origen.has_style:
        destino._style = copy(origen._style)
    if origen.number_format:
        destino.number_format = origen.number_format
    if origen.alignment:
        destino.alignment = copy(origen.alignment)
    if origen.border:
        destino.border = copy(origen.border)
    if origen.fill:
        destino.fill = copy(origen.fill)
    if origen.font:
        destino.font = copy(origen.font)
    if origen.protection:
        destino.protection = copy(origen.protection)

# Copia el ancho de una columna a otra
def copiar_estilo_columna(ws, col_origen, col_destino):
    letra_origen = get_column_letter(col_origen)
    letra_destino = get_column_letter(col_destino)

    if ws.column_dimensions[letra_origen].width:
        ws.column_dimensions[letra_destino].width = ws.column_dimensions[letra_origen].width

# Busca una fila donde el texto sea exactamente igual
def buscar_fila_por_texto(ws, texto, col=2, fila_inicio=1, fila_fin=None):
    objetivo = normalizar_texto(texto)
    fila_fin = fila_fin or ws.max_row

    for fila in range(fila_inicio, fila_fin + 1):
        valor = ws.cell(row=fila, column=col).value
        if normalizar_texto(valor) == objetivo:
            return fila

    return None

# Busca una fila que contenga cierto texto
def buscar_fila_contiene_texto(ws, texto, col=2, fila_inicio=1, fila_fin=None):
    objetivo = normalizar_texto(texto)
    fila_fin = fila_fin or ws.max_row

    for fila in range(fila_inicio, fila_fin + 1):
        valor = ws.cell(row=fila, column=col).value
        if objetivo in normalizar_texto(valor):
            return fila

    return None

# Busca donde esta la columna Total dentro de los cuadros
def buscar_columna_total(ws, fila_encabezado, col_inicio=3):
    for col in range(col_inicio, ws.max_column + 1):
        if normalizar_texto(ws.cell(row=fila_encabezado, column=col).value) == "TOTAL":
            return col
    return None

# Aplica formato del monto
def aplicar_formato_monto(celda):
    celda.number_format = FORMATO_MONTO_REPORTE

# Aplica formato de porcentaje
def aplicar_formato_porcentaje(celda):
    celda.number_format = FORMATO_PORCENTAJE_REPORTE

# Quita el fondo de una celda
def quitar_relleno(celda):
    """Deja la celda sin color de fondo."""
    celda.fill = PatternFill(fill_type=None)

# Aplica el formato final de cada bloque
def aplicar_estilo_final_bloque(
    ws,
    titulo_bloque,
    col_inicio,
    col_total,
    fila_total,
    filas_credito,
    fila_acumulado,
):
    
    fill_detalle = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL_DETALLE)
    fill_total = PatternFill(fill_type="solid", fgColor=COLOR_TOTAL_FILA)

    col_validacion = col_total + 1
    letra_total = get_column_letter(col_total)

    # Celdas de detalle en la columna Total.
    for fila in filas_credito:
        celda = ws.cell(row=fila, column=col_total)
        celda.fill = copy(fill_detalle)
        celda.number_format = FORMATO_MONTO_REPORTE

    # Fila Total: el color llega solo hasta la columna Total.
    for col in range(col_inicio, col_total + 1):
        celda = ws.cell(row=fila_total, column=col)
        celda.fill = copy(fill_total)
        celda.number_format = FORMATO_MONTO_REPORTE

    # Celda validadora al costado de Total, sin color.
    celda_validador = ws.cell(row=fila_total, column=col_validacion)
    quitar_relleno(celda_validador)
    celda_validador.number_format = "General"

    titulo_normalizado = normalizar_texto(titulo_bloque)

    if "REAL" in titulo_normalizado:
        celda_control = "$G$21"
    elif "PROYECTADA" in titulo_normalizado:
        celda_control = "$H$21"
    elif "MOROSA" in titulo_normalizado:
        celda_control = "$I$21"
    else:
        celda_control = None

    if celda_control:
        celda_validador.value = f'=IF({celda_control}={letra_total}{fila_total},"OK","REVISAR")'
    else:
        celda_validador.value = None

    # La fila Acumulado no debe colorear la columna Total ni el validador.
    if fila_acumulado:
        for col in range(col_total, col_validacion + 1):
            celda = ws.cell(row=fila_acumulado, column=col)
            celda.value = None
            quitar_relleno(celda)
            celda.number_format = "General"


# =========================================================
# 7. ESCRIBIR BASE OCULTA
# =========================================================

# Crea la hoja, pega la base de datos y la oculta
def escribir_base_en_reporte(wb, base):
    if NOMBRE_HOJA_BASE in wb.sheetnames:
        wb.remove(wb[NOMBRE_HOJA_BASE])

    ws = wb.create_sheet(NOMBRE_HOJA_BASE)

    for fila in dataframe_to_rows(base, index=False, header=True):
        ws.append(fila)

    encabezados = {celda.value: celda.column for celda in ws[1]}

    for columna in COLUMNAS_FECHA:
        if columna in encabezados:
            col_idx = encabezados[columna]
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=col_idx).number_format = "dd/mm/yyyy"

    for columna in encabezados:
        if es_columna_numero(columna):
            col_idx = encabezados[columna]
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=col_idx).number_format = "#,##0.00"

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.freeze_panes = "A2"
    ws.sheet_state = "hidden"

    return ws


# =========================================================
# 8. FÓRMULAS DE EXCEL
# =========================================================

# Detecta las columnas tipo_inmueble
def mapa_columnas_excel(base):
    return {
        columna: get_column_letter(indice + 1)
        for indice, columna in enumerate(base.columns)
    }

# Arma una part de formula para detectar si una fila tiene departamento, estacionamiento, deposito o local
def rango_base(columna, columnas_excel, max_fila):
    letra = columnas_excel[columna]
    return f"'{NOMBRE_HOJA_BASE}'!${letra}$2:${letra}${max_fila}"

# Arma la formula para identificar el tipo de financiamiento
def formula_palabra_financiamiento(celda_tipo):
    return (
        f'IF(ISNUMBER(SEARCH("DIRECTO",UPPER({celda_tipo}))),"DIRECTO",'
        f'IF(ISNUMBER(SEARCH("HIPOT",UPPER({celda_tipo}))),"HIPOT",'
        f'IF(ISNUMBER(SEARCH("AHORRO",UPPER({celda_tipo}))),"AHORRO",UPPER({celda_tipo}))))'
    )

# Arma la condición del estado: Pendiente, Cancelado y Vencido
def condicion_estado(tipo_estado, rng_estado):
    tipo_estado = normalizar_texto(tipo_estado)

    if tipo_estado == "PENDIENTE":
        return f'--ISNUMBER(SEARCH("PENDIENTE",UPPER(TRIM({rng_estado}))))'

    if tipo_estado == "VENCIDO":
        return f'--ISNUMBER(SEARCH("VENCIDO",UPPER(TRIM({rng_estado}))))'

    if tipo_estado == "CANCELADO":
        return f'--ISNUMBER(SEARCH("CANCELADO",UPPER(TRIM({rng_estado}))))'

    return "1"

# Arma la formula de las cobranzas por periodo
def formula_periodo_cobranza(
    celda_proyecto,
    celda_tipo_credito,
    celda_periodo,
    columna_monto,
    columna_fecha,
    tipo_estado,
    columnas_excel,
    max_fila,
):
    rng_proyecto = rango_base("Proyecto", columnas_excel, max_fila)
    rng_estado = rango_base("Estado", columnas_excel, max_fila)
    rng_fecha = rango_base(columna_fecha, columnas_excel, max_fila)
    rng_monto = rango_base(columna_monto, columnas_excel, max_fila)
    rng_financiamiento = rango_base("TipoFinanciamiento", columnas_excel, max_fila)

    palabra_financiamiento = formula_palabra_financiamiento(celda_tipo_credito)

    fecha_inicio = f'IF({celda_periodo}<3000,DATE({celda_periodo},1,1),{celda_periodo})'
    fecha_fin = f'IF({celda_periodo}<3000,DATE({celda_periodo}+1,1,1),EDATE({celda_periodo},1))'

    condiciones = [
        f'--ISNUMBER(SEARCH({celda_proyecto},{rng_proyecto}))',
        f'--({rng_fecha}>={fecha_inicio})',
        f'--({rng_fecha}<{fecha_fin})',
        f'--ISNUMBER(SEARCH({palabra_financiamiento},UPPER({rng_financiamiento})))',
    ]

    if normalizar_texto(tipo_estado) == "REAL":
        condiciones.append(f'--({rng_monto}>0)')
    else:
        condiciones.append(condicion_estado(tipo_estado, rng_estado))

    return (
        f'=IFERROR(SUMPRODUCT('
        + ','.join(condiciones)
        + f',{rng_monto}),0)'
    )


# =========================================================
# 9. CLASIFICACIÓN DE INMUEBLES PARA FÓRMULAS
# =========================================================

# Detecta la columna TipoInmueble
def obtener_columnas_tipo_inmueble(columnas_excel):
    columnas_tipo = []

    for columna in columnas_excel:
        if re.match(r"^TipoInmueble_(\d+)$", str(columna)):
            columnas_tipo.append(columna)

    columnas_tipo = sorted(
        columnas_tipo,
        key=lambda x: int(re.search(r"_(\d+)$", x).group(1)),
    )

    return columnas_tipo

# Arma una parte de la formula para detectar si una fila tiene departamento, estacionamiento, deposito o local
def formula_suma_busqueda_tipo(columnas_tipo, columnas_excel, max_fila, palabra):
    partes = []

    for columna in columnas_tipo:
        rng_tipo = rango_base(columna, columnas_excel, max_fila)
        partes.append(f'--ISNUMBER(SEARCH("{palabra}",UPPER({rng_tipo})))')

    if not partes:
        return "(0)"

    return "(" + "+".join(partes) + ")"

# Arma la formula para calcular montos por tipo de inmueble
def formula_tipo_inmueble(
    celda_proyecto,
    columna_monto,
    tipo_estado,
    tipo_reporte,
    columnas_excel,
    max_fila,
):
    rng_proyecto = rango_base("Proyecto", columnas_excel, max_fila)
    rng_estado = rango_base("Estado", columnas_excel, max_fila)
    rng_monto = rango_base(columna_monto, columnas_excel, max_fila)

    columnas_tipo = obtener_columnas_tipo_inmueble(columnas_excel)

    tiene_departamento = formula_suma_busqueda_tipo(columnas_tipo, columnas_excel, max_fila, "DEPART")
    tiene_estacionamiento = formula_suma_busqueda_tipo(columnas_tipo, columnas_excel, max_fila, "ESTACION")
    tiene_deposito = formula_suma_busqueda_tipo(columnas_tipo, columnas_excel, max_fila, "DEPOS")
    tiene_local = formula_suma_busqueda_tipo(columnas_tipo, columnas_excel, max_fila, "LOCAL")
    tiene_comercial = formula_suma_busqueda_tipo(columnas_tipo, columnas_excel, max_fila, "COMERCIAL")

    tipo_reporte = normalizar_texto(tipo_reporte)

    if tipo_reporte == "DEPARTAMENTO":
        condicion_tipo = (
            f'--({tiene_departamento}>0),'
            f'--({tiene_estacionamiento}=0)'
        )

    elif tipo_reporte == "DEPARTAMENTO + ESTACIONAMIENTO":
        condicion_tipo = (
            f'--({tiene_departamento}>0),'
            f'--({tiene_estacionamiento}>0)'
        )

    elif tipo_reporte == "ESTACIONAMIENTOS":
        condicion_tipo = (
            f'--({tiene_departamento}=0),'
            f'--({tiene_estacionamiento}>0)'
        )

    elif tipo_reporte == "DEPOSITOS":
        condicion_tipo = (
            f'--({tiene_departamento}=0),'
            f'--({tiene_estacionamiento}=0),'
            f'--({tiene_deposito}>0)'
        )

    elif tipo_reporte == "LOCAL COMERCIAL":
        condicion_tipo = f'--(({tiene_local}+{tiene_comercial})>0)'

    else:
        return "=0"

    condiciones = [
        f'--ISNUMBER(SEARCH({celda_proyecto},{rng_proyecto}))',
        condicion_tipo,
    ]

    if normalizar_texto(tipo_estado) == "REAL":
        condiciones.append(f'--({rng_monto}>0)')
    else:
        condiciones.append(condicion_estado(tipo_estado, rng_estado))

    return (
        f'=IFERROR(SUMPRODUCT('
        + ','.join(condiciones)
        + f',{rng_monto}),0)'
    )


# =========================================================
# 10. CÁLCULOS EN PYTHON: VENDIDO, UNIDADES Y MOROSOS
# =========================================================

# Detecta que indices de inmuebles tiene 
def obtener_indices_inmueble(base):
    indices = set()

    for columna in base.columns:
        coincidencia = re.match(r"^TipoInmueble_(\d+)$", str(columna))
        if coincidencia:
            indices.add(int(coincidencia.group(1)))

    return sorted(indices)

# Clasifica si una fila es departamento, dpt + estc, estc, dep o local 
def clasificar_tipos_fila(tipos):
    tipos_norm = [normalizar_texto(t) for t in tipos if str(t).strip() not in ["", "nan", "None"]]

    tiene_departamento = any("DEPART" in t for t in tipos_norm)
    tiene_estacionamiento = any("ESTACION" in t for t in tipos_norm)
    tiene_deposito = any("DEPOS" in t for t in tipos_norm)
    tiene_local = any(("LOCAL" in t or "COMERCIAL" in t) for t in tipos_norm)

    if tiene_departamento and tiene_estacionamiento:
        return "Departamento + Estacionamiento"
    if tiene_departamento:
        return "Departamento"
    if tiene_estacionamiento:
        return "Estacionamientos"
    if tiene_deposito:
        return "Depositos"
    if tiene_local:
        return "Local Comercial"

    return None

# Valida si un dato sirve para identificar a un cliente
def valor_cliente_valido(valor):
    if pd.isna(valor):
        return False

    texto = str(valor).strip()

    if texto == "":
        return False

    texto_norm = normalizar_texto(texto)

    valores_invalidos = {
        "-",
        "--",
        "0",
        "00",
        "00000000",
        "000000000",
        "00000000000",
        "S/D",
        "SD",
        "SIN DNI",
        "SIN DOCUMENTO",
        "SIN DOC",
        "NO TIENE",
        "NINGUNO",
        "NAN",
        "NONE",
        "NULL",
    }

    if texto_norm in valores_invalidos:
        return False

    if re.fullmatch(r"0+", texto_norm):
        return False

    return True

# Obtiene el identificador del cliente para contar morosos
def obtener_id_cliente(row):
    """
    Prioriza el nombre del titular y su documento
    """

    # Primero usar el nombre visible del cliente.
    columnas_nombre = [
        "Nombres_Titular",
        "Cliente",
    ]

    for col in columnas_nombre:
        if col in row.index:
            valor = row.get(col)

            if valor_cliente_valido(valor):
                return normalizar_texto(valor)

    # Respaldo: usar documento solo si es un valor real.
    columnas_documento = [
        "NroDocumento",
        "Nro_Documento",
        "Documento",
        "NumeroDocumento",
        "NumDocumento",
        "DNI",
    ]

    for col in columnas_documento:
        if col in row.index:
            valor = row.get(col)

            if valor_cliente_valido(valor):
                return normalizar_texto(valor)

    return None

# Calcula precio vendido y la cantidad de unidades por tipo de inmueble
def calcular_vendido_unidades(base, proyecto):
    base_temp = base.copy()
    base_temp["Proyecto_Corto"] = base_temp["Proyecto"].apply(obtener_nombre_proyecto)
    base_proyecto = base_temp[base_temp["Proyecto_Corto"] == proyecto].copy()

    indices = obtener_indices_inmueble(base_proyecto)

    resultado = {
        "Departamento": {"monto": 0.0, "unidades": 0, "vistos": set()},
        "Departamento + Estacionamiento": {"monto": 0.0, "unidades": 0, "vistos": set()},
        "Estacionamientos": {"monto": 0.0, "unidades": 0, "vistos": set()},
        "Depositos": {"monto": 0.0, "unidades": 0, "vistos": set()},
        "Local Comercial": {"monto": 0.0, "unidades": 0, "vistos": set()},
    }

    for _, row in base_proyecto.iterrows():
        tipos = [row.get(f"TipoInmueble_{idx}", "") for idx in indices]
        categoria = clasificar_tipos_fila(tipos)

        if categoria is None or categoria not in resultado:
            continue

        for idx in indices:
            tipo = row.get(f"TipoInmueble_{idx}", "")
            nro = row.get(f"NroInmueble_{idx}", "")

            if pd.isna(tipo) or str(tipo).strip() == "":
                continue

            if pd.isna(nro) or str(nro).strip() == "":
                nro = f"SIN_NRO_{idx}_{row.name}"

            clave = f"{idx}|{normalizar_texto(tipo)}|{normalizar_texto(nro)}"

            if clave in resultado[categoria]["vistos"]:
                continue

            precio = row.get(f"PrecioVenta_{idx}", 0)
            precio = pd.to_numeric(precio, errors="coerce")
            precio = 0 if pd.isna(precio) else float(precio)

            resultado[categoria]["monto"] += precio
            resultado[categoria]["unidades"] += 1
            resultado[categoria]["vistos"].add(clave)

    resumen = {}

    for categoria, data in resultado.items():
        resumen[categoria] = {
            "monto": data["monto"],
            "unidades": data["unidades"],
        }

    resumen["Departamentos"] = {
        "monto": resumen["Departamento"]["monto"] + resumen["Departamento + Estacionamiento"]["monto"],
        "unidades": resumen["Departamento"]["unidades"] + resumen["Departamento + Estacionamiento"]["unidades"],
    }

    resumen["Total"] = {
        "monto": (
            resumen["Departamentos"]["monto"]
            + resumen["Estacionamientos"]["monto"]
            + resumen["Depositos"]["monto"]
            + resumen["Local Comercial"]["monto"]
        ),
        "unidades": (
            resumen["Departamentos"]["unidades"]
            + resumen["Estacionamientos"]["unidades"]
            + resumen["Depositos"]["unidades"]
            + resumen["Local Comercial"]["unidades"]
        ),
    }

    return resumen

# Calcula clientes morosos y los días promedio de atraso por tipo de inmueble
def calcular_morosos_y_atraso(base, proyecto):
    fecha_corte = obtener_fecha_corte_reporte()

    if fecha_corte is None:
        fecha_corte = pd.Timestamp.today().normalize()

    base_temp = base.copy()
    base_temp["Proyecto_Corto"] = base_temp["Proyecto"].apply(obtener_nombre_proyecto)
    base_proyecto = base_temp[base_temp["Proyecto_Corto"] == proyecto].copy()

    estado = base_proyecto["Estado"].astype(str).apply(normalizar_texto)
    estado_vencido = estado.str.contains("VENCIDO", na=False)

    base_morosa = base_proyecto[
        estado_vencido
        & (base_proyecto["Fecha_Programada"].notna())
        & (base_proyecto["Fecha_Programada"] <= fecha_corte)
        & (base_proyecto["SaldoPorPagarCuota"].fillna(0) > 0)
    ].copy()

    categorias = [
        "Departamento",
        "Departamento + Estacionamiento",
        "Estacionamientos",
        "Depositos",
        "Local Comercial",
        "Departamentos",
        "Total",
    ]

    resultado = {
        categoria: {"clientes": 0, "dias": 0}
        for categoria in categorias
    }

    if base_morosa.empty:
        return resultado

    indices = obtener_indices_inmueble(base_morosa)
    registros = []

    for _, row in base_morosa.iterrows():
        tipos = [row.get(f"TipoInmueble_{idx}", "") for idx in indices]
        categoria = clasificar_tipos_fila(tipos)

        if categoria is None:
            continue

        cliente = obtener_id_cliente(row)

        if cliente is None:
            continue

        saldo = pd.to_numeric(row.get("SaldoPorPagarCuota", 0), errors="coerce")
        saldo = 0 if pd.isna(saldo) else float(saldo)

        if saldo <= 0:
            continue

        fecha_programada = pd.to_datetime(row.get("Fecha_Programada"), errors="coerce")

        if pd.isna(fecha_programada):
            continue

        dias = max((fecha_corte - fecha_programada).days, 0)

        registros.append({
            "categoria": categoria,
            "cliente": cliente,
            "saldo": saldo,
            "dias": dias,
        })

    if not registros:
        return resultado

    df = pd.DataFrame(registros)

    def calcular_para(cats):
        temp = df[df["categoria"].isin(cats)].copy()

        if temp.empty:
            return {"clientes": 0, "dias": 0}

        deuda_cliente = temp.groupby("cliente", as_index=False)["saldo"].sum()

        clientes_con_deuda = deuda_cliente[
            deuda_cliente["saldo"] > 0
        ]["cliente"].tolist()

        if not clientes_con_deuda:
            return {"clientes": 0, "dias": 0}

        temp_deuda = temp[temp["cliente"].isin(clientes_con_deuda)].copy()
        dias_promedio = temp_deuda[temp_deuda["saldo"] > 0]["dias"].mean()

        if pd.isna(dias_promedio):
            dias_promedio = 0

        return {
            "clientes": len(set(clientes_con_deuda)),
            "dias": int(round(dias_promedio, 0)),
        }

    resultado["Departamento"] = calcular_para(["Departamento"])
    resultado["Departamento + Estacionamiento"] = calcular_para(["Departamento + Estacionamiento"])
    resultado["Estacionamientos"] = calcular_para(["Estacionamientos"])
    resultado["Depositos"] = calcular_para(["Depositos"])
    resultado["Local Comercial"] = calcular_para(["Local Comercial"])

    resultado["Departamentos"] = calcular_para([
        "Departamento",
        "Departamento + Estacionamiento",
    ])

    resultado["Total"] = calcular_para([
        "Departamento",
        "Departamento + Estacionamiento",
        "Estacionamientos",
        "Depositos",
        "Local Comercial",
    ])

    return resultado


# =========================================================
# 11. ACTUALIZACIÓN DEL CUADRO SUPERIOR
# =========================================================

# Llena el cuadro con la informacion
def actualizar_resumen_superior(ws, base, proyecto, columnas_excel, max_fila):
    celda_proyecto = "$B$6"

    vendido = calcular_vendido_unidades(base, proyecto)
    morosos = calcular_morosos_y_atraso(base, proyecto)

    filas = {
        "Departamentos": buscar_fila_por_texto(ws, "Departamentos", fila_inicio=1, fila_fin=25),
        "Departamento": buscar_fila_por_texto(ws, "Departamento", fila_inicio=1, fila_fin=25),
        "Departamento + Estacionamiento": buscar_fila_por_texto(ws, "Departamento + Estacionamiento", fila_inicio=1, fila_fin=25),
        "Estacionamientos": buscar_fila_por_texto(ws, "Estacionamientos", fila_inicio=1, fila_fin=25),
        "Depositos": buscar_fila_por_texto(ws, "Depositos", fila_inicio=1, fila_fin=25),
        "Local Comercial": buscar_fila_por_texto(ws, "Local Comercial", fila_inicio=1, fila_fin=25),
        "Total": buscar_fila_por_texto(ws, "Total", fila_inicio=1, fila_fin=25),
    }

    detalles = [
        "Departamento",
        "Departamento + Estacionamiento",
        "Estacionamientos",
        "Depositos",
        "Local Comercial",
    ]

    for nombre in detalles:
        fila = filas.get(nombre)
        if not fila:
            continue

        ws.cell(row=fila, column=3).value = vendido.get(nombre, {}).get("monto", 0)
        ws.cell(row=fila, column=4).value = vendido.get(nombre, {}).get("unidades", 0)

        ws.cell(row=fila, column=6).value = f"=IF(C{fila}=0,0,G{fila}/C{fila})"

        ws.cell(row=fila, column=7).value = formula_tipo_inmueble(
            celda_proyecto=celda_proyecto,
            columna_monto="Monto_Cuota_Pagado",
            tipo_estado="REAL",
            tipo_reporte=nombre,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
        )

        ws.cell(row=fila, column=8).value = formula_tipo_inmueble(
            celda_proyecto=celda_proyecto,
            columna_monto="SaldoPorPagarCuota",
            tipo_estado="PENDIENTE",
            tipo_reporte=nombre,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
        )

        ws.cell(row=fila, column=9).value = formula_tipo_inmueble(
            celda_proyecto=celda_proyecto,
            columna_monto="SaldoPorPagarCuota",
            tipo_estado="VENCIDO",
            tipo_reporte=nombre,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
        )

        ws.cell(row=fila, column=10).value = f"=IF(C{fila}=0,0,I{fila}/C{fila})"
        ws.cell(row=fila, column=11).value = morosos.get(nombre, {}).get("clientes", 0)
        ws.cell(row=fila, column=12).value = morosos.get(nombre, {}).get("dias", 0)

    # Filas resumen calculadas directamente para evitar duplicidades en clientes morosos.
    for nombre in ["Departamentos", "Total"]:
        fila = filas.get(nombre)
        if fila:
            ws.cell(row=fila, column=3).value = vendido.get(nombre, {}).get("monto", 0)
            ws.cell(row=fila, column=4).value = vendido.get(nombre, {}).get("unidades", 0)
            ws.cell(row=fila, column=11).value = morosos.get(nombre, {}).get("clientes", 0)
            ws.cell(row=fila, column=12).value = morosos.get(nombre, {}).get("dias", 0)

    fila_departamentos = filas.get("Departamentos")
    fila_departamento = filas.get("Departamento")
    fila_dep_est = filas.get("Departamento + Estacionamiento")

    if fila_departamentos and fila_departamento and fila_dep_est:
        for col in [7, 8, 9]:
            letra = get_column_letter(col)
            ws.cell(row=fila_departamentos, column=col).value = f"=SUM({letra}{fila_departamento}:{letra}{fila_dep_est})"

        ws.cell(row=fila_departamentos, column=6).value = f"=IF(C{fila_departamentos}=0,0,G{fila_departamentos}/C{fila_departamentos})"
        ws.cell(row=fila_departamentos, column=10).value = f"=IF(C{fila_departamentos}=0,0,I{fila_departamentos}/C{fila_departamentos})"

    fila_total = filas.get("Total")
    filas_sumar = [
        filas.get("Departamentos"),
        filas.get("Estacionamientos"),
        filas.get("Depositos"),
        filas.get("Local Comercial"),
    ]
    filas_sumar = [fila for fila in filas_sumar if fila]

    if fila_total and filas_sumar:
        for col in [7, 8, 9]:
            letra = get_column_letter(col)
            partes = [f"{letra}{fila}" for fila in filas_sumar]
            ws.cell(row=fila_total, column=col).value = "=" + "+".join(partes)

        ws.cell(row=fila_total, column=6).value = f"=IF(C{fila_total}=0,0,G{fila_total}/C{fila_total})"
        ws.cell(row=fila_total, column=10).value = f"=IF(C{fila_total}=0,0,I{fila_total}/C{fila_total})"

    # Formatos del cuadro superior.
    for fila in range(1, min(ws.max_row, 25) + 1):
        for col in [3, 4, 7, 8, 9, 11, 12]:
            ws.cell(row=fila, column=col).number_format = FORMATO_MONTO_REPORTE

        ws.cell(row=fila, column=6).number_format = FORMATO_PORCENTAJE_REPORTE
        ws.cell(row=fila, column=10).number_format = FORMATO_PORCENTAJE_REPORTE


# =========================================================
# 12. ACTUALIZACIÓN DE BLOQUES DE COBRANZA
# =========================================================

# Actualiza los cuadros de cobranza real, proyectada y morosa
def actualizar_bloque_cobranzas(
    ws,
    titulo_bloque,
    periodos,
    columnas_excel,
    max_fila,
    columna_monto,
    columna_fecha,
    tipo_estado,
):
    fila_titulo = buscar_fila_por_texto(ws, titulo_bloque)

    if not fila_titulo:
        raise ValueError(f"No se encontró el bloque: {titulo_bloque} en la hoja {ws.title}")

    fila_encabezado = buscar_fila_contiene_texto(
        ws,
        "Tipo",
        fila_inicio=fila_titulo,
        fila_fin=min(fila_titulo + 5, ws.max_row),
    )

    if not fila_encabezado:
        raise ValueError(f"No se encontró la fila de encabezados del bloque: {titulo_bloque}")

    fila_total = buscar_fila_por_texto(
        ws,
        "Total",
        fila_inicio=fila_encabezado + 1,
        fila_fin=min(fila_encabezado + 30, ws.max_row),
    )

    if not fila_total:
        raise ValueError(f"No se encontró la fila Total del bloque: {titulo_bloque}")

    fila_acumulado = buscar_fila_por_texto(
        ws,
        "Acumulado",
        fila_inicio=fila_total + 1,
        fila_fin=min(fila_total + 10, ws.max_row),
    )

    filas_credito = []
    for fila in range(fila_encabezado + 1, fila_total):
        valor = ws.cell(row=fila, column=2).value
        if valor not in [None, ""]:
            filas_credito.append(fila)

    col_inicio = 3
    total_col_anterior = buscar_columna_total(ws, fila_encabezado, col_inicio=col_inicio)

    if not total_col_anterior:
        total_col_anterior = min(ws.max_column, col_inicio + 10)

    cantidad_periodos = len(periodos)
    col_total = col_inicio + cantidad_periodos
    col_limpiar = max(ws.max_column, total_col_anterior + 10, col_total + 10)

    filas_bloque = [fila_encabezado] + filas_credito + [fila_total]
    if fila_acumulado:
        filas_bloque.append(fila_acumulado)

    celda_limpia_referencia = ws.cell(row=1, column=1)

    # Limpiar valores antiguos.
    for fila in filas_bloque:
        for col in range(col_inicio, col_limpiar + 1):
            ws.cell(row=fila, column=col).value = None

    # Copiar estilos de periodos.
    for col in range(col_inicio, col_total):
        if col < total_col_anterior:
            col_estilo = col
        else:
            col_estilo = max(total_col_anterior - 1, col_inicio)

        copiar_estilo_columna(ws, col_estilo, col)

        for fila in filas_bloque:
            copiar_estilo(ws.cell(row=fila, column=col_estilo), ws.cell(row=fila, column=col))

    # Copiar estilo especial de columna Total.
    copiar_estilo_columna(ws, total_col_anterior, col_total)
    for fila in filas_bloque:
        copiar_estilo(ws.cell(row=fila, column=total_col_anterior), ws.cell(row=fila, column=col_total))

    # Limpiar estilos sobrantes a la derecha del Total.
    for fila in filas_bloque:
        for col in range(col_total + 1, col_limpiar + 1):
            celda = ws.cell(row=fila, column=col)
            celda.value = None
            celda._style = copy(celda_limpia_referencia._style)
            celda.number_format = "General"

    # Encabezados de periodos.
    for i, periodo in enumerate(periodos):
        col = col_inicio + i
        celda = ws.cell(row=fila_encabezado, column=col)

        if periodo["tipo"] == "anio":
            celda.value = int(periodo["valor"])
            celda.number_format = "0"
        else:
            celda.value = periodo["valor"].to_pydatetime()
            celda.number_format = "mmm-yy"

    # Encabezado Total.
    celda_total_header = ws.cell(row=fila_encabezado, column=col_total)
    celda_total_header.value = "Total"

    if cantidad_periodos > 0:
        celda_encabezado_referencia = ws.cell(row=fila_encabezado, column=col_total - 1)
    else:
        celda_encabezado_referencia = ws.cell(row=fila_encabezado, column=total_col_anterior)

    copiar_estilo(celda_encabezado_referencia, celda_total_header)

    # Fórmulas por tipo de crédito.
    for fila in filas_credito:
        for i in range(cantidad_periodos):
            col = col_inicio + i
            letra_col = get_column_letter(col)

            ws.cell(row=fila, column=col).value = formula_periodo_cobranza(
                celda_proyecto="$B$6",
                celda_tipo_credito=f"$B{fila}",
                celda_periodo=f"{letra_col}${fila_encabezado}",
                columna_monto=columna_monto,
                columna_fecha=columna_fecha,
                tipo_estado=tipo_estado,
                columnas_excel=columnas_excel,
                max_fila=max_fila,
            )
            ws.cell(row=fila, column=col).number_format = FORMATO_MONTO_REPORTE

        letra_inicio = get_column_letter(col_inicio)
        letra_fin = get_column_letter(col_total - 1)

        if cantidad_periodos > 0:
            ws.cell(row=fila, column=col_total).value = f"=SUM({letra_inicio}{fila}:{letra_fin}{fila})"
        else:
            ws.cell(row=fila, column=col_total).value = "=0"

        ws.cell(row=fila, column=col_total).number_format = FORMATO_MONTO_REPORTE

    # Total por periodo.
    for i in range(cantidad_periodos):
        col = col_inicio + i
        letra_col = get_column_letter(col)

        if filas_credito:
            ws.cell(row=fila_total, column=col).value = f"=SUM({letra_col}{min(filas_credito)}:{letra_col}{max(filas_credito)})"
        else:
            ws.cell(row=fila_total, column=col).value = "=0"

        ws.cell(row=fila_total, column=col).number_format = FORMATO_MONTO_REPORTE

    # Total general del bloque.
    if cantidad_periodos > 0:
        letra_inicio = get_column_letter(col_inicio)
        letra_fin = get_column_letter(col_total - 1)
        ws.cell(row=fila_total, column=col_total).value = f"=SUM({letra_inicio}{fila_total}:{letra_fin}{fila_total})"
    else:
        ws.cell(row=fila_total, column=col_total).value = "=0"

    ws.cell(row=fila_total, column=col_total).number_format = FORMATO_MONTO_REPORTE

    # Acumulado.
    if fila_acumulado:
        for i in range(cantidad_periodos):
            col = col_inicio + i
            letra_col = get_column_letter(col)

            if i == 0:
                ws.cell(row=fila_acumulado, column=col).value = f"={letra_col}{fila_total}"
            else:
                letra_anterior = get_column_letter(col - 1)
                ws.cell(row=fila_acumulado, column=col).value = f"={letra_anterior}{fila_acumulado}+{letra_col}{fila_total}"

            ws.cell(row=fila_acumulado, column=col).number_format = FORMATO_MONTO_REPORTE

        ws.cell(row=fila_acumulado, column=col_total).value = None
        quitar_relleno(ws.cell(row=fila_acumulado, column=col_total))
        ws.cell(row=fila_acumulado, column=col_total).number_format = "General"

    aplicar_estilo_final_bloque(
        ws=ws,
        titulo_bloque=titulo_bloque,
        col_inicio=col_inicio,
        col_total=col_total,
        fila_total=fila_total,
        filas_credito=filas_credito,
        fila_acumulado=fila_acumulado,
    )


# =========================================================
# 13. GENERAR REPORTE
# =========================================================

# Genera el archivo final
def generar_reporte_cobranzas(base):
    validar_columnas_reporte(base)

    ruta_plantilla = buscar_plantilla_reporte()
    ruta_reporte = SALIDA_LISTA / NOMBRE_REPORTE_COBRANZAS

    wb = load_workbook(ruta_plantilla)

    if NOMBRE_HOJA_PLANTILLA in wb.sheetnames:
        ws_plantilla = wb[NOMBRE_HOJA_PLANTILLA]
    else:
        hojas_visibles = [ws for ws in wb.worksheets if ws.title != NOMBRE_HOJA_BASE]
        if not hojas_visibles:
            raise ValueError("La plantilla no tiene una hoja visible para usar como modelo.")
        ws_plantilla = hojas_visibles[0]

    # Escribir base ajustada dentro del reporte y ocultarla.
    escribir_base_en_reporte(wb, base)

    proyectos = obtener_proyectos_reporte(base)
    print("📌 Proyectos para Reporte_Cobranzas:", ", ".join(proyectos))

    nombre_modelo = "__Modelo_Cobranzas__"

    if nombre_modelo in wb.sheetnames:
        wb.remove(wb[nombre_modelo])

    ws_modelo = wb.copy_worksheet(ws_plantilla)
    ws_modelo.title = nombre_modelo
    ws_modelo.sheet_state = "hidden"

    # Dejar solo modelo y base.
    for ws in list(wb.worksheets):
        if ws.title not in [nombre_modelo, NOMBRE_HOJA_BASE]:
            wb.remove(ws)

    columnas_excel = mapa_columnas_excel(base)
    max_fila = len(base) + 1

    for proyecto in proyectos:
        nombre_hoja = nombre_hoja_valido(proyecto)

        ws = wb.copy_worksheet(ws_modelo)
        ws.title = nombre_hoja
        ws.sheet_state = "visible"
        ws["B6"] = proyecto

        periodos_real, periodos_proy, periodos_morosa = obtener_periodos_proyecto(base, proyecto)

        actualizar_resumen_superior(
            ws=ws,
            base=base,
            proyecto=proyecto,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
        )

        actualizar_bloque_cobranzas(
            ws=ws,
            titulo_bloque="COBRANZA REAL",
            periodos=periodos_real,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
            columna_monto="Monto_Cuota_Pagado",
            columna_fecha="FechaPago",
            tipo_estado="REAL",
        )

        actualizar_bloque_cobranzas(
            ws=ws,
            titulo_bloque="COBRANZA PROYECTADA",
            periodos=periodos_proy,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
            columna_monto="SaldoPorPagarCuota",
            columna_fecha="Fecha_Programada",
            tipo_estado="PENDIENTE",
        )

        actualizar_bloque_cobranzas(
            ws=ws,
            titulo_bloque="COBRANZA MOROSA",
            periodos=periodos_morosa,
            columnas_excel=columnas_excel,
            max_fila=max_fila,
            columna_monto="SaldoPorPagarCuota",
            columna_fecha="Fecha_Programada",
            tipo_estado="VENCIDO",
        )

        print(f"✅ Hoja creada: {ws.title}")

    if nombre_modelo in wb.sheetnames:
        wb.remove(wb[nombre_modelo])

    if NOMBRE_HOJA_BASE in wb.sheetnames:
        wb[NOMBRE_HOJA_BASE].sheet_state = "hidden"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    wb.save(ruta_reporte)
    return ruta_reporte


# =========================================================
# 14. EJECUCIÓN
# =========================================================

# Decide si primero ejecuta descarga y luego carga la base ajustada
def preparar_base_cobranzas(actualizar_base=True):
    if actualizar_base:
        ejecutar_descarga_cobranzas()

    base_consolidada, ruta_base = cargar_base_consolidada()
    return base_consolidada, ruta_base

# Genera solo el reporte usando la base cargada
def generar_solo_reporte_cobranzas(base_consolidada):
    """
    Genera Reporte_Cobranzas.xlsx usando directamente Base_Ajustada.
    """
    ruta_reporte = generar_reporte_cobranzas(base_consolidada)
    return ruta_reporte

# Ejecuta el flujo completo: base + reporte
def ejecutar_reporte_cobranzas(actualizar_base=True):
    base_consolidada, ruta_base = preparar_base_cobranzas(actualizar_base=actualizar_base)
    ruta_reporte = generar_solo_reporte_cobranzas(base_consolidada)

    print("====================================")
    print("REPORTE DE COBRANZAS TERMINADO")
    print("====================================")
    print("✅ Base ajustada utilizada:")
    print(f"   {ruta_base}")
    print("✅ Reporte cobranzas generado con éxito:")
    print(f"   {ruta_reporte}")

    return base_consolidada, ruta_base, ruta_reporte

# Funcion de compatibilidad con nombres anteriores
def consolidar_cobranzas():
    """Compatibilidad con nombres anteriores."""
    return ejecutar_reporte_cobranzas(actualizar_base=EJECUTAR_DESCARGA)

# Punto de entrada cuando ejecutas el archivo directamente
def main():
    ejecutar_reporte_cobranzas(actualizar_base=EJECUTAR_DESCARGA)


if __name__ == "__main__":
    main()
